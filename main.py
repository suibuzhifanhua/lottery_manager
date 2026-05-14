"""
彩票历史数据管理系统
功能：查询/导入/导出/在线抓取/购彩账本
支持：双色球（SSQ）/ 大乐透（DLT）
"""

import os
import sqlite3
import sys
import threading

# ─── PyTorch 优化禁用（打包后用户名问题）─────────────────────────
os.environ["TORCHINDUCTOR_DISABLE"] = "1"
os.environ["PYTORCH_JIT"] = "0"
os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:512"
# 打包后可能没有 USER 环境变量，导致 PyTorch dynamo 失败
if not os.environ.get("USER"):
    os.environ["USER"] = os.environ.get("USERNAME", "default_user")
if not os.environ.get("USERNAME"):
    os.environ["USERNAME"] = os.environ.get("USER", "default_user")

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date

# ─── 版本信息 ─────────────────────────────────────────────────
__VERSION__ = "2.4.0"
__BUILD_DATE__ = "2026-04-30"

# ─── 版本更新日志 ─────────────────────────────────────────────
VERSION_HISTORY = [
    {
        "version": "2.4.0",
        "date": "2026-04-30",
        "type": "improve",
        "title": "🚀 启动速度优化",
        "changes": [
            "【删除】移除图片 OCR 识别功能模块",
            "【优化】精简打包体积，移除 rapidocr_onnxruntime 依赖",
            "【优化】模块延迟加载，启动速度提升",
        ],
        "highlights": "启动更快，体积更小"
    },
    {
        "version": "2.3.3",
        "date": "2026-04-30",
        "type": "improve",
        "title": "🔍 OCR 多注解析优化",
        "changes": [
            '【优化】新增 "红球-蓝球" 格式专用解析（识别 XX-XX 分隔符）',
            "【优化】支持带序号（①②③④⑤）的多注识别",
            "【优化】每注一行逐行解析，避免多注混在一起识别错误",
            "【优化】蓝球识别优先取末尾数字（1-16），大幅提升准确率",
        ],
        "highlights": "5注彩票完整识别，蓝球准确率提升"
    },
    {
        "version": "2.3.2",
        "date": "2026-04-30",
        "type": "bugfix",
        "title": "🔧 OCR 引擎迁移修复",
        "changes": [
            "【修复】exe 上传图片报错 No module named 'torch._dynamo.polyfills.fx'",
            "【修复】将 OCR 引擎从 cnocr（torch）迁移到 rapidocr_onnxruntime（纯 onnx）",
            "【修复】打包不再包含 torch（约 200MB 体积减少）",
            "【修复】spec 文件移除 torch/cnocr/cnstd hiddenimports",
            "【修复】环境变量修复逻辑提前到所有 import 之前",
        ],
        "highlights": "彻底解决打包后 OCR 无法使用的问题，exe 体积更小"
    },
    {
        "version": "2.3.1",
        "date": "2026-04-30",
        "type": "improve",
        "title": "🔍 图片识别优化",
        "changes": [
            "【优化】图片识别支持多注一次识别",
            "【优化】新增图片预处理（对比度/锐化/缩放），提升识别率",
            "【优化】识别结果支持手动逐注编辑号码",
            "【优化】新增左侧图片预览，右侧可查看 OCR 原始文本",
            "【优化】多分辨率尝试提升复杂图片识别效果",
            "【修复】导出功能支持大乐透（使用正确列名）",
            "【修复】修复 USERNAME 未设置导致 OCR 引擎初始化失败的问题",
        ],
        "highlights": "识别准确率大幅提升，支持多注批量录入"
    },
    {
        "version": "2.3.0",
        "date": "2026-04-29",
        "type": "feature",
        "title": "📷 图片识别功能",
        "changes": [
            "【新增】上传彩票图片自动识别彩票类型（双色球/大乐透）",
            "【新增】自动识别期号和购买号码",
            "【新增】一键添加到购彩账本",
        ],
        "highlights": "支持拍照上传彩票图片，快速录入账本"
    },
    {
        "version": "2.2.1",
        "date": "2026-04-21",
        "type": "bugfix",
        "title": "🔧 随机森林/XGBoost Bug 修复",
        "changes": [
            "【修复】随机森林 index 越界（feat_row 索引错误）",
            "【修复】XGBoost DLL 未正确打包",
            "【修复】spec 文件 XGBoost 依赖配置",
        ],
        "highlights": "解决随机森林和XGBoost完全无法使用的问题"
    },
    {
        "version": "2.2.0",
        "date": "2026-04-21",
        "type": "bugfix",
        "title": "🔧 核心Bug修复",
        "changes": [
            "【修复】XGBoost 预测报错 XGBoostLibraryNotFound",
            "【修复】随机森林预测报错 index 33 越界",
            "【优化】数据库批量插入提速 20-50x",
            "【优化】随机森林特征预计算提速 10-50x",
            "【优化】XGBoost 并行训练提速 5-8x",
        ],
        "highlights": "解决预测模块完全无法使用的问题"
    },
    {
        "version": "2.1.0",
        "date": "2026-04-20",
        "type": "feature",
        "title": "✨ 功能增强",
        "changes": [
            "【新增】12种预测方案并行执行",
            "【新增】多方案并行预测",
            "【优化】启动速度优化",
            "【新增】购彩账本支持大乐透",
        ],
        "highlights": "预测速度大幅提升"
    },
]

# ❌ 旧代码（启动时就加载重量级库，导致启动慢）：
# import predict
# import predict_dlt

# ✅ 新代码：延迟导入，仅在首次调用预测功能时才加载
_predict = None
_predict_dlt = None

def _load_predict_modules():
    """延迟加载预测模块（sklearn/xgboost等重量级库）"""
    global _predict, _predict_dlt
    if _predict is None:
        import predict
        import predict_dlt
        _predict = predict
        _predict_dlt = predict_dlt
    return _predict, _predict_dlt

def get_predict_module():
    """获取 predict 模块（延迟加载）"""
    p, _ = _load_predict_modules()
    return p

def get_predict_dlt_module():
    """获取 predict_dlt 模块（延迟加载）"""
    _, pd = _load_predict_modules()
    return pd

# ─── 模块延迟加载（启动优化）────────────────────────────────────
# 这些模块只在需要时才加载，大幅提升启动速度
_fetch_data = None
_fetch_dlt = None
_ledger = None
lg = None  # ledger 模块别名（用于兼容性）

def _ensure_fetch_data():
    """确保 fetch_data 模块已加载"""
    global _fetch_data
    if _fetch_data is None:
        import fetch_data
        _fetch_data = fetch_data
    return _fetch_data

def _ensure_fetch_dlt():
    """确保 fetch_dlt 模块已加载"""
    global _fetch_dlt
    if _fetch_dlt is None:
        import fetch_dlt
        _fetch_dlt = fetch_dlt
    return _fetch_dlt

def _ensure_ledger():
    """确保 ledger 模块已加载"""
    global _ledger, lg
    if _ledger is None:
        import ledger
        _ledger = ledger
        lg = ledger  # 保持兼容性
    return _ledger
try:
    from tkcalendar import DateEntry as _TkDateEntry
    _HAS_TKCALENDAR = True
except ImportError:
    _HAS_TKCALENDAR = False

# ─── 路径配置 ───────────────────────────────────────────────
def get_base_dir():
    """获取应用基础目录（PyInstaller 打包后的临时解压目录，用于加载数据文件）"""
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

def get_app_dir():
    """获取 exe 所在目录（用于写入输出文件）"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_dir()
APP_DIR = get_app_dir()
# 数据库放在 exe 所在目录，这样打包后也能找到
DB_PATH = os.path.join(APP_DIR, "ssq.db")

# ─── 颜色常量 ───────────────────────────────────────────────
RED_BG   = "#EE4444"
BLUE_BG  = "#3366CC"
WHITE_FG = "#FFFFFF"
BG_MAIN  = "#F5F5F5"
BG_CARD  = "#FFFFFF"
ACCENT   = "#E53935"

# ─── 数据库 ─────────────────────────────────────────────────
def init_db():
    fd = _ensure_fetch_data()
    fdl = _ensure_fetch_dlt()
    ledger_module = _ensure_ledger()
    fd.init_db(DB_PATH)
    fdl.init_dlt_db(DB_PATH)
    ledger_module.init_ledger_db(DB_PATH)
    _init_version_db()


def _init_version_db():
    """初始化版本更新日志表"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # app_config 表（用于存储配置信息）
    c.execute("""
        CREATE TABLE IF NOT EXISTS app_config (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    
    # 版本更新日志表
    c.execute("""
        CREATE TABLE IF NOT EXISTS app_version_log (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            version       TEXT NOT NULL UNIQUE,
            build_date    TEXT,
            update_type   TEXT,
            title         TEXT,
            changes       TEXT,
            highlights    TEXT,
            created_at    TEXT DEFAULT (datetime('now','localtime'))
        )
    """)
    
    # 插入版本更新日志
    for v in VERSION_HISTORY:
        c.execute("""
            INSERT OR IGNORE INTO app_version_log 
            (version, build_date, update_type, title, changes, highlights)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (v["version"], v["date"], v["type"], v["title"], 
              "\n".join(v["changes"]), v["highlights"]))
    
    conn.commit()
    conn.close()


def get_version_log():
    """获取版本更新日志"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT version, build_date, update_type, title, changes, highlights FROM app_version_log ORDER BY version DESC")
    rows = c.fetchall()
    conn.close()
    return rows


def get_current_version():
    """获取当前版本"""
    return __VERSION__

def query_db(sql, params=()):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(sql, params)
    rows = c.fetchall()
    conn.close()
    return rows

def exec_db(sql, params=()):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(sql, params)
    conn.commit()
    conn.close()

def upsert_record(rec):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT OR REPLACE INTO ssq_history
        (issue,draw_date,red1,red2,red3,red4,red5,red6,blue,jackpot,prize1_count,prize2_count,sales)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, rec)
    conn.commit()
    conn.close()

def count_records(lottery_type="ssq"):
    table = "ssq_history" if lottery_type == "ssq" else "dlt_history"
    try:
        r = query_db(f"SELECT COUNT(*) FROM {table}")
        return r[0][0] if r else 0
    except Exception:
        return 0

# ─── Excel 导入导出 ─────────────────────────────────────────
COLUMNS = ["期号","开奖日期","红球1","红球2","红球3","红球4","红球5","红球6","蓝球","奖池金额","一等奖注数","二等奖注数","销售额"]
COLUMNS_DLT = ["期号","开奖日期","前区1","前区2","前区3","前区4","前区5","后区1","后区2","奖池金额","一等奖注数","二等奖注数","销售额"]

def _import_openpyxl():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import sys
    m = sys.modules[__name__]
    m.openpyxl = openpyxl
    m.Font = Font
    m.PatternFill = PatternFill
    m.Alignment = Alignment
    m.Border = Border
    m.Side = Side
    m.get_column_letter = get_column_letter

def export_template(path, lottery_type="ssq"):
    """导出Excel模板
    
    Args:
        path: 保存路径
        lottery_type: 'ssq' 双色球 或 'dlt' 大乐透
    """
    _import_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if lottery_type == "dlt":
        ws.title = "大乐透数据模板"
        columns = COLUMNS_DLT
        # 大乐透示例：前区5个 + 后区2个
        example = ["26001","2026-01-01",5,9,14,19,24,8,12,"1000000000",5,30,"300000000"]
    else:
        ws.title = "双色球数据模板"
        columns = COLUMNS
        # 双色球示例：红球6个 + 蓝球1个
        example = ["26001","2026-01-01",5,12,18,23,29,33,11,"100000000",3,20,"200000000"]
    
    # 写入表头
    ws.append(columns)
    header_row = ws[1]
    for i, cell in enumerate(header_row):
        col_idx = i + 1
        if lottery_type == "dlt":
            # 大乐透：前区(2-6)红，后区(7-8)蓝，期号日期深灰
            if 2 <= col_idx <= 6:
                cell.fill = PatternFill("solid", fgColor="CC0000")
            elif col_idx in (7, 8):
                cell.fill = PatternFill("solid", fgColor="0033CC")
            else:
                cell.fill = PatternFill("solid", fgColor="333333")
        else:
            # 双色球：红球(3-8)红，蓝球(9)蓝，期号日期深灰
            if 3 <= col_idx <= 8:
                cell.fill = PatternFill("solid", fgColor="CC0000")
            elif col_idx == 9:
                cell.fill = PatternFill("solid", fgColor="0033CC")
            else:
                cell.fill = PatternFill("solid", fgColor="333333")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    
    # 写入示例行
    ws.append(example)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, cell in enumerate(ws[2], 1):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        if lottery_type == "dlt":
            if col_idx in (7, 8):  # 后区
                cell.fill = PatternFill("solid", fgColor="EEF0FF")
                cell.font = Font(color="0033CC", bold=True)
            elif 2 <= col_idx <= 6:  # 前区
                cell.fill = PatternFill("solid", fgColor="FFEEEE")
                cell.font = Font(color="CC0000", bold=True)
        else:
            if 3 <= col_idx <= 8:  # 红球
                cell.fill = PatternFill("solid", fgColor="FFEEEE")
                cell.font = Font(color="CC0000", bold=True)
            elif col_idx == 9:  # 蓝球
                cell.fill = PatternFill("solid", fgColor="EEF0FF")
                cell.font = Font(color="0033CC", bold=True)
    
    _set_col_width(ws)
    wb.save(path)

def export_data(path, rows, lottery_type="ssq"):
    """导出数据到Excel，支持双色球(ssq)和大乐透(dlt)
    
    Args:
        path: 保存路径
        rows: 数据行（数据库原始数据）
        lottery_type: 'ssq' 双色球 或 'dlt' 大乐透
    """
    _import_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 根据彩票类型设置工作表标题和列名
    if lottery_type == "dlt":
        ws.title = "大乐透历史数据"
        columns = COLUMNS_DLT
        # 大乐透列样式：前区红色，后区蓝色
        def _ball_style(cell, col_idx):
            if 2 <= col_idx <= 6:  # 前区 1-5
                cell.fill = PatternFill("solid", fgColor="FFEEEE")
                cell.font = Font(color="CC0000", bold=True)
            elif col_idx in (7, 8):  # 后区 1-2
                cell.fill = PatternFill("solid", fgColor="EEF0FF")
                cell.font = Font(color="0033CC", bold=True)
    else:
        ws.title = "双色球历史数据"
        columns = COLUMNS
        # 双色球列样式：红球1-6红色，蓝球蓝色
        def _ball_style(cell, col_idx):
            if 3 <= col_idx <= 8:  # 红球 1-6
                cell.fill = PatternFill("solid", fgColor="FFEEEE")
                cell.font = Font(color="CC0000", bold=True)
            elif col_idx == 9:  # 蓝球
                cell.fill = PatternFill("solid", fgColor="EEF0FF")
                cell.font = Font(color="0033CC", bold=True)
    
    # 写入表头
    ws.append(columns)
    header_row = ws[1]
    for i, cell in enumerate(header_row):
        col_idx = i + 1
        if lottery_type == "dlt":
            # 大乐透：前区(2-6)红，后区(7-8)蓝，期号日期深灰
            if 2 <= col_idx <= 6:
                cell.fill = PatternFill("solid", fgColor="CC0000")
            elif col_idx in (7, 8):
                cell.fill = PatternFill("solid", fgColor="0033CC")
            else:
                cell.fill = PatternFill("solid", fgColor="333333")
        else:
            # 双色球：红球(3-8)红，蓝球(9)蓝，期号日期深灰
            if 3 <= col_idx <= 8:
                cell.fill = PatternFill("solid", fgColor="CC0000")
            elif col_idx == 9:
                cell.fill = PatternFill("solid", fgColor="0033CC")
            else:
                cell.fill = PatternFill("solid", fgColor="333333")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    
    # 写入数据行
    for row_idx, row in enumerate(rows, 2):
        ws.append(list(row))
        thin = Side(style='thin', color='DDDDDD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            # 球号列特殊样式
            if lottery_type == "dlt":
                if col_idx in (7, 8):  # 后区
                    cell.fill = PatternFill("solid", fgColor="EEF0FF")
                    cell.font = Font(color="0033CC", bold=True)
                elif 2 <= col_idx <= 6:  # 前区
                    cell.fill = PatternFill("solid", fgColor="FFEEEE")
                    cell.font = Font(color="CC0000", bold=True)
                else:
                    bg = "FFFFFF" if row_idx % 2 == 0 else "F8F8F8"
                    cell.fill = PatternFill("solid", fgColor=bg)
            else:
                if 3 <= col_idx <= 8:  # 红球
                    cell.fill = PatternFill("solid", fgColor="FFEEEE")
                    cell.font = Font(color="CC0000", bold=True)
                elif col_idx == 9:  # 蓝球
                    cell.fill = PatternFill("solid", fgColor="EEF0FF")
                    cell.font = Font(color="0033CC", bold=True)
                else:
                    bg = "FFFFFF" if row_idx % 2 == 0 else "F8F8F8"
                    cell.fill = PatternFill("solid", fgColor=bg)
    
    _set_col_width(ws)
    wb.save(path)

def _write_header(ws):
    ws.append(COLUMNS)
    header_row = ws[1]
    for i, cell in enumerate(header_row):
        if i < 9:  # 期号到蓝球
            cell.fill = PatternFill("solid", fgColor="CC0000" if i > 1 else "333333")
        else:
            cell.fill = PatternFill("solid", fgColor="1A3A6B")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

def _style_data_rows(ws, start, end):
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx in range(start, end + 1):
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            # 红球列(3-8)着色
            if 3 <= col_idx <= 8:
                cell.fill = PatternFill("solid", fgColor="FFEEEE")
                cell.font = Font(color="CC0000", bold=True)
            # 蓝球列(9)着色
            elif col_idx == 9:
                cell.fill = PatternFill("solid", fgColor="EEF0FF")
                cell.font = Font(color="0033CC", bold=True)
            else:
                bg = "FFFFFF" if row_idx % 2 == 0 else "F8F8F8"
                cell.fill = PatternFill("solid", fgColor=bg)

def _set_col_width(ws):
    widths = [10, 12, 7, 7, 7, 7, 7, 7, 7, 16, 12, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def import_excel(path):
    _import_openpyxl()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    inserted = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        try:
            issue = str(row[0]).strip().zfill(5)
            draw_date = str(row[1]).strip() if row[1] else ""
            reds = [int(row[i]) for i in range(2, 8)]
            blue = int(row[8])
            jackpot = str(row[9]) if row[9] else ""
            p1 = int(row[10]) if row[10] else 0
            p2 = int(row[11]) if row[11] else 0
            sales = str(row[12]) if len(row) > 12 and row[12] else ""
            upsert_record((issue, draw_date, *reds, blue, jackpot, p1, p2, sales))
            inserted += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {e}")
    return inserted, errors

# ─── 主窗口 ─────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"🎱 彩票历史数据管理系统  v{__VERSION__}")
        self.geometry("1200x780")
        self.minsize(900, 600)
        self.configure(bg=BG_MAIN)
        # 当前彩票类型：'ssq' = 双色球，'dlt' = 大乐透
        self._lottery_type = tk.StringVar(value="ssq")

        # ── 启动闪屏：先显示窗口，再异步初始化 ──
        splash = tk.Label(self, text="🎱 正在初始化彩票管理系统...",
                         bg=BG_MAIN, fg="#888888",
                         font=("Microsoft YaHei", 14))
        splash.place(relx=0.5, rely=0.5, anchor="center")

        self._build_ui()

        # 延迟执行耗时操作，让窗口先渲染出来
        self.after(100, self._deferred_init)

    def _deferred_init(self):
        """延迟初始化：数据库 + 数据加载，完成后隐藏闪屏"""
        try:
            init_db()
            self._load_data()
            # 检查版本更新
            self._check_version_update()
        except Exception as e:
            print(f"[启动错误] {e}")
        finally:
            # 销毁闪屏（如果还在显示）
            for w in self.winfo_children():
                if isinstance(w, tk.Label) and w.cget("text").startswith("🎱 正在初始化"):
                    w.destroy()
                    break

    def _check_version_update(self):
        """检查版本更新，显示更新日志"""
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            # 获取上次记录的版本
            c.execute("SELECT value FROM app_config WHERE key='last_version'")
            row = c.fetchone()
            last_version = row[0] if row else None
            conn.close()

            # 如果版本有更新，显示更新日志
            if last_version != __VERSION__:
                self.after(500, lambda: self._show_version_update_popup(last_version))
                # 更新记录
                exec_db("INSERT OR REPLACE INTO app_config (key, value) VALUES ('last_version', ?)", (__VERSION__,))
        except Exception:
            # app_config 表可能不存在，忽略
            pass

    def _show_version_update_popup(self, old_version):
        """显示版本更新弹窗"""
        win = tk.Toplevel(self)
        win.title("🎉 版本更新")
        win.geometry("500x400")
        win.resizable(False, False)
        win.configure(bg="#1A237E")
        win.grab_set()

        # 标题
        tk.Label(win, text="🎉",
                bg="#1A237E", fg="#FFD700",
                font=("Microsoft YaHei", 40)).pack(pady=(20, 10))
        tk.Label(win, text=f"已更新至 v{__VERSION__}",
                bg="#1A237E", fg="white",
                font=("Microsoft YaHei", 18, "bold")).pack()
        if old_version:
            tk.Label(win, text=f"（从 v{old_version} 升级）",
                    bg="#1A237E", fg="#90CAF9",
                    font=("Microsoft YaHei", 10)).pack(pady=(5, 15))
        else:
            tk.Label(win, text="（首次运行）",
                    bg="#1A237E", fg="#90CAF9",
                    font=("Microsoft YaHei", 10)).pack(pady=(5, 15))

        # 获取最新版本的更新日志
        logs = get_version_log()
        if logs:
            latest = logs[0]
            changes = latest[4] if latest[4] else ""

            content_frame = tk.Frame(win, bg="#283593", padx=20, pady=15)
            content_frame.pack(fill="both", expand=True, padx=30, pady=10)

            tk.Label(content_frame, text=latest[3] or "本次更新",
                    bg="#283593", fg="#FFD54F",
                    font=("Microsoft YaHei", 11, "bold")).pack(anchor="w")

            for line in (changes.split("\n") if changes else []):
                if line.strip():
                    tk.Label(content_frame, text=line,
                            bg="#283593", fg="#E8EAF6",
                            font=("Microsoft YaHei", 9),
                            anchor="w").pack(fill="x", pady=2)

            if latest[5]:
                tk.Label(content_frame, text=f"⭐ 重点：{latest[5]}",
                        bg="#1A237E", fg="#FFEB3B",
                        font=("Microsoft YaHei", 9, "bold"),
                        padx=10, pady=5).pack(pady=(10, 0))

        # 按钮
        btn_frame = tk.Frame(win, bg="#1A237E")
        btn_frame.pack(pady=15)
        tk.Button(btn_frame, text="查看完整更新日志 →", 
                 command=lambda: [win.destroy(), self._show_about()],
                 bg="#3949AB", fg="white",
                 font=("Microsoft YaHei", 10),
                 padx=20, pady=8, relief="flat",
                 cursor="hand2").pack(side="left", padx=5)
        tk.Button(btn_frame, text="知道了", command=win.destroy,
                 bg="#757575", fg="white",
                 font=("Microsoft YaHei", 10),
                 padx=20, pady=8, relief="flat",
                 cursor="hand2").pack(side="left", padx=5)

    # ── UI构建 ──────────────────────────────────────────────
    def _build_ui(self):
        self._build_toolbar()
        self._build_filter()
        self._build_table()
        self._build_statusbar()

    # ── 工具栏悬停效果辅助 ─────────────────────────────────────
    @staticmethod
    def _bind_hover(btn, normal_bg, hover_bg):
        btn.bind("<Enter>", lambda e: btn.configure(bg=hover_bg))
        btn.bind("<Leave>", lambda e: btn.configure(bg=normal_bg))

    def _build_toolbar(self):
        # ── 整体容器：深红渐变感背景 ────────────────────────────
        bar = tk.Frame(self, bg="#B71C1C", pady=0)
        bar.pack(fill="x")

        # ── 左侧品牌区 ─────────────────────────────────────────
        brand = tk.Frame(bar, bg="#B71C1C")
        brand.pack(side="left", fill="y")
        tk.Label(brand, text="🎱",
                 bg="#B71C1C", fg="white",
                 font=("Segoe UI Emoji", 20)).pack(side="left", padx=(14, 4), pady=8)
        title_box = tk.Frame(brand, bg="#B71C1C")
        title_box.pack(side="left", pady=6)
        self._title_lbl = tk.Label(title_box, text="双色球历史数据管理系统",
                 bg="#B71C1C", fg="white",
                 font=("Microsoft YaHei", 13, "bold"))
        self._title_lbl.pack(anchor="w")
        self._subtitle_lbl = tk.Label(title_box, text="SSQ History Manager",
                 bg="#B71C1C", fg="#FFCDD2",
                 font=("Microsoft YaHei", 8))
        self._subtitle_lbl.pack(anchor="w")

        # ── 垂直分隔线 ─────────────────────────────────────────
        tk.Frame(bar, bg="#D32F2F", width=1).pack(side="left", fill="y", padx=10, pady=6)

        # ── 彩票类型切换区 ────────────────────────────────────
        type_area = tk.Frame(bar, bg="#B71C1C")
        type_area.pack(side="left", fill="y", pady=0)
        tk.Label(type_area, text="彩票类型", bg="#B71C1C", fg="#FFCDD2",
                 font=("Microsoft YaHei", 7)).pack(anchor="center")
        type_btns = tk.Frame(type_area, bg="#B71C1C")
        type_btns.pack()

        def make_type_btn(parent, text, val, active_bg, inactive_bg):
            """彩票类型切换按钮"""
            font_obj = ("Microsoft YaHei", 9, "bold")
            tmp = tk.Label(parent, text=text, font=font_obj)
            tmp.update_idletasks()
            tw = tmp.winfo_reqwidth()
            tmp.destroy()
            w = tw + 24
            h = 30
            r = 6
            cv = tk.Canvas(parent, width=w, height=h,
                           bg="#B71C1C", highlightthickness=0, cursor="hand2")

            def draw():
                is_active = (self._lottery_type.get() == val)
                color = active_bg if is_active else inactive_bg
                border_color = "#FFFF88" if is_active else color
                cv.delete("all")
                cv.create_arc(0, 0, 2*r, 2*r, start=90, extent=90, fill=color, outline=color)
                cv.create_arc(w-2*r, 0, w, 2*r, start=0, extent=90, fill=color, outline=color)
                cv.create_arc(0, h-2*r, 2*r, h, start=180, extent=90, fill=color, outline=color)
                cv.create_arc(w-2*r, h-2*r, w, h, start=270, extent=90, fill=color, outline=color)
                cv.create_rectangle(r, 0, w-r, h, fill=color, outline=color)
                cv.create_rectangle(0, r, w, h-r, fill=color, outline=color)
                txt_color = "#FFFF00" if is_active else "#FFFFFF"
                cv.create_text(w//2, h//2, text=text, fill=txt_color,
                               font=font_obj, anchor="center")

            draw()

            def on_click(e):
                self._lottery_type.set(val)
                self._on_lottery_type_change()
                # 重绘两个按钮
                for cv2, draw2 in _type_btn_refs:
                    draw2()

            cv.bind("<Button-1>", on_click)
            return cv, draw

        _type_btn_refs = []
        ssq_cv, ssq_draw = make_type_btn(type_btns, "🎱 双色球", "ssq", "#C62828", "#7B1010")
        ssq_cv.pack(side="left", padx=2)
        _type_btn_refs.append((ssq_cv, ssq_draw))

        dlt_cv, dlt_draw = make_type_btn(type_btns, "🌟 大乐透", "dlt", "#1565C0", "#0A3570")
        dlt_cv.pack(side="left", padx=2)
        _type_btn_refs.append((dlt_cv, dlt_draw))

        # 保存引用以便后续重绘
        self._type_btn_refs = _type_btn_refs

        # ── 垂直分隔线 ─────────────────────────────────────────
        tk.Frame(bar, bg="#D32F2F", width=1).pack(side="left", fill="y", padx=10, pady=6)

        # ── 右侧按钮区（Canvas 实现圆角胶囊按钮）──────────────
        btn_area = tk.Frame(bar, bg="#B71C1C")
        btn_area.pack(side="right", padx=10, pady=0, fill="y")

        def make_capsule_btn(parent, text, cmd, bg="#C62828", hover="#E53935", width=None):
            """在 Canvas 上绘制圆角胶囊按钮，带悬停高亮"""
            font_obj = ("Microsoft YaHei", 9)
            # 计算文字宽度
            tmp = tk.Label(parent, text=text, font=font_obj)
            tmp.update_idletasks()
            tw = tmp.winfo_reqwidth()
            tmp.destroy()
            w = (width or tw) + 24
            h = 30
            r = 6  # 圆角半径

            cv = tk.Canvas(parent, width=w, height=h,
                           bg="#B71C1C", highlightthickness=0, cursor="hand2")

            def draw(color):
                cv.delete("all")
                # 圆角矩形（用 8 段弧拼出）
                cv.create_arc(0, 0, 2*r, 2*r, start=90, extent=90, fill=color, outline=color)
                cv.create_arc(w-2*r, 0, w, 2*r, start=0, extent=90, fill=color, outline=color)
                cv.create_arc(0, h-2*r, 2*r, h, start=180, extent=90, fill=color, outline=color)
                cv.create_arc(w-2*r, h-2*r, w, h, start=270, extent=90, fill=color, outline=color)
                cv.create_rectangle(r, 0, w-r, h, fill=color, outline=color)
                cv.create_rectangle(0, r, w, h-r, fill=color, outline=color)
                cv.create_text(w//2, h//2, text=text, fill="white",
                               font=font_obj, anchor="center")

            draw(bg)
            cv.bind("<Enter>",  lambda e: draw(hover))
            cv.bind("<Leave>",  lambda e: draw(bg))
            cv.bind("<Button-1>", lambda e: cmd())
            return cv

        def make_sep(parent):
            tk.Frame(parent, bg="#D32F2F", width=1).pack(side="left", fill="y", padx=6, pady=8)

        # ── 分组1：数据录入 ────────────────────────────────────
        grp1 = tk.Frame(btn_area, bg="#B71C1C")
        grp1.pack(side="left", fill="y", pady=7)
        tk.Label(grp1, text="数据", bg="#B71C1C", fg="#FFCDD2",
                 font=("Microsoft YaHei", 7)).pack(anchor="center")
        grp1_btns = tk.Frame(grp1, bg="#B71C1C")
        grp1_btns.pack()
        for txt, cmd in [("➕ 录入", self._manual_add),
                          ("📥 导入", self._import),
                          ("📋 模板", self._download_template),
                          ("📤 导出", self._export)]:
            make_capsule_btn(grp1_btns, txt, cmd).pack(side="left", padx=2)

        make_sep(btn_area)

        # ── 分组2：工具 ────────────────────────────────────────
        grp2 = tk.Frame(btn_area, bg="#B71C1C")
        grp2.pack(side="left", fill="y", pady=7)
        tk.Label(grp2, text="工具", bg="#B71C1C", fg="#FFCDD2",
                 font=("Microsoft YaHei", 7)).pack(anchor="center")
        grp2_btns = tk.Frame(grp2, bg="#B71C1C")
        grp2_btns.pack()
        make_capsule_btn(grp2_btns, "📒 购彩账本", self._show_ledger,
                         bg="#2E7D32", hover="#388E3C").pack(side="left", padx=2)
        make_capsule_btn(grp2_btns, "🔮 预测下一期", self._show_prediction,
                         bg="#4527A0", hover="#512DA8").pack(side="left", padx=2)

        make_sep(btn_area)

        # ── 分组3：数据更新 ────────────────────────────────────
        grp3 = tk.Frame(btn_area, bg="#B71C1C")
        grp3.pack(side="left", fill="y", pady=7)
        tk.Label(grp3, text="更新", bg="#B71C1C", fg="#FFCDD2",
                 font=("Microsoft YaHei", 7)).pack(anchor="center")
        grp3_btns = tk.Frame(grp3, bg="#B71C1C")
        grp3_btns.pack()
        make_capsule_btn(grp3_btns, "🔄 增量更新", self._fetch_recent,
                         bg="#1565C0", hover="#1976D2").pack(side="left", padx=2)
        make_capsule_btn(grp3_btns, "⬇ 全量抓取", self._fetch_all,
                         bg="#1565C0", hover="#1976D2").pack(side="left", padx=2)

        # ── 分组4：关于（靠右）───────────────────────────────────
        grp4 = tk.Frame(btn_area, bg="#B71C1C")
        grp4.pack(side="right", fill="y", pady=7, padx=(10, 14))
        tk.Label(grp4, text=f"v{__VERSION__}", bg="#B71C1C", fg="#FFCDD2",
                 font=("Microsoft YaHei", 7)).pack(anchor="center")
        grp4_btns = tk.Frame(grp4, bg="#B71C1C")
        grp4_btns.pack()
        make_capsule_btn(grp4_btns, "ℹ️ 关于", self._show_about,
                         bg="#455A64", hover="#607D8B").pack(side="left", padx=2)

    def _show_about(self):
        """显示关于对话框，包含版本信息和更新日志"""
        win = tk.Toplevel(self)
        win.title("ℹ️ 关于")
        win.geometry("580x600")
        win.resizable(False, False)
        win.configure(bg="#1A1A2E")
        win.grab_set()

        # 标题区域
        hdr = tk.Frame(win, bg="#16213E", pady=15)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🎱 彩票历史数据管理系统",
                bg="#16213E", fg="white",
                font=("Microsoft YaHei", 16, "bold")).pack()
        tk.Label(hdr, text=f"版本 {__VERSION__}  |  构建日期 {__BUILD_DATE__}",
                bg="#16213E", fg="#64B5F6",
                font=("Microsoft YaHei", 10)).pack(pady=(5, 0))

        # 更新日志区域
        canvas = tk.Canvas(win, bg="#1A1A2E", highlightthickness=0)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=15, pady=10)

        content = tk.Frame(canvas, bg="#1A1A2E")
        cw = canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # 获取更新日志
        logs = get_version_log()

        for idx, (version, build_date, update_type, title, changes, highlights) in enumerate(logs):
            # 版本卡片
            card = tk.Frame(content, bg="#16213E", padx=15, pady=12)
            card.pack(fill="x", pady=(0, 10))

            # 版本标题行
            type_colors = {"bugfix": "#E53935", "feature": "#43A047", "optimize": "#1E88E5"}
            type_color = type_colors.get(update_type, "#9E9E9E")

            row = tk.Frame(card, bg="#16213E")
            row.pack(fill="x")
            tk.Label(row, text=f"v{version}",
                    bg="#16213E", fg="#FFD700",
                    font=("Microsoft YaHei", 12, "bold")).pack(side="left")
            tk.Label(row, text=build_date,
                    bg="#16213E", fg="#9E9E9E",
                    font=("Microsoft YaHei", 9)).pack(side="left", padx=10)
            tk.Label(row, text=update_type.upper() if update_type else "",
                    bg=type_color, fg="white",
                    font=("Microsoft YaHei", 8, "bold")).pack(side="right")

            tk.Label(card, text=title,
                    bg="#16213E", fg="white",
                    font=("Microsoft YaHei", 10, "bold"),
                    anchor="w").pack(fill="x", pady=(8, 4))

            # 改动内容
            if changes:
                for line in changes.split("\n"):
                    if line.strip():
                        tk.Label(card, text=line,
                                bg="#16213E", fg="#B0BEC5",
                                font=("Microsoft YaHei", 9),
                                anchor="w").pack(fill="x")

            # 重点说明
            if highlights:
                hl_frame = tk.Frame(card, bg="#263238", padx=8, pady=6)
                hl_frame.pack(fill="x", pady=(8, 0))
                tk.Label(hl_frame, text=f"⭐ {highlights}",
                        bg="#263238", fg="#FFD54F",
                        font=("Microsoft YaHei", 9)).pack(anchor="w")

        # 关闭按钮
        btn_frame = tk.Frame(win, bg="#1A1A2E", pady=10)
        btn_frame.pack()
        tk.Button(btn_frame, text="关闭", command=win.destroy,
                 bg="#455A64", fg="white",
                 font=("Microsoft YaHei", 10),
                 padx=30, pady=5, relief="flat",
                 cursor="hand2").pack()

    def _on_lottery_type_change(self):
        """切换彩票类型时更新标题、表格列等"""
        lt = self._lottery_type.get()
        if lt == "ssq":
            self._title_lbl.configure(text="双色球历史数据管理系统")
            self._subtitle_lbl.configure(text="SSQ History Manager")
        else:
            self._title_lbl.configure(text="大乐透历史数据管理系统")
            self._subtitle_lbl.configure(text="DLT History Manager")
        # 重建表格列（双色球和大乐透列结构不同）
        self._rebuild_table_columns()
        self._load_data()
        # 若预测窗口已打开，关闭它（因为类型变了）
        if hasattr(self, "_pred_win") and self._pred_win and self._pred_win.winfo_exists():
            self._pred_win.destroy()
            self._pred_win = None

    def _build_filter(self):
        frame = tk.Frame(self, bg=BG_CARD, pady=8, padx=10,
                         highlightbackground="#DDDDDD", highlightthickness=1)
        frame.pack(fill="x", padx=10, pady=(6, 0))

        lbl = dict(bg=BG_CARD, font=("Microsoft YaHei", 9))
        ent = dict(font=("Microsoft YaHei", 9), relief="solid", bd=1)

        tk.Label(frame, text="期号范围：", **lbl).pack(side="left")
        self.var_issue_from = tk.StringVar()
        self.var_issue_to   = tk.StringVar()
        tk.Entry(frame, textvariable=self.var_issue_from, width=8, **ent).pack(side="left")
        tk.Label(frame, text=" ~ ", **lbl).pack(side="left")
        tk.Entry(frame, textvariable=self.var_issue_to, width=8, **ent).pack(side="left", padx=(0,16))

        tk.Label(frame, text="日期范围：", **lbl).pack(side="left")
        self.var_date_from = tk.StringVar()
        self.var_date_to   = tk.StringVar()
        tk.Entry(frame, textvariable=self.var_date_from, width=11, **ent).pack(side="left")
        tk.Label(frame, text=" ~ ", **lbl).pack(side="left")
        tk.Entry(frame, textvariable=self.var_date_to, width=11, **ent).pack(side="left", padx=(0,16))

        tk.Label(frame, text="包含红球：", **lbl).pack(side="left")
        self.var_red = tk.StringVar()
        tk.Entry(frame, textvariable=self.var_red, width=8, **ent).pack(side="left", padx=(0,6))
        # 蓝球标签随彩票类型变化
        self._blue_filter_lbl = tk.Label(frame, text="蓝球：", **lbl)
        self._blue_filter_lbl.pack(side="left")
        self.var_blue = tk.StringVar()
        tk.Entry(frame, textvariable=self.var_blue, width=5, **ent).pack(side="left", padx=(0,16))

        tk.Button(frame, text="🔍 查询", command=self._load_data,
                  bg=ACCENT, fg="white", relief="flat",
                  font=("Microsoft YaHei", 9, "bold"),
                  padx=12, pady=3, cursor="hand2").pack(side="left", padx=4)
        tk.Button(frame, text="↺ 重置", command=self._reset_filter,
                  bg="#888", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9),
                  padx=10, pady=3, cursor="hand2").pack(side="left", padx=4)

    def _build_table(self):
        self._table_frame = tk.Frame(self, bg=BG_MAIN)
        self._table_frame.pack(fill="both", expand=True, padx=10, pady=8)

        self._build_tree_in_frame(self._table_frame)

    def _build_tree_in_frame(self, frame):
        """在 frame 中创建 Treeview（支持双色球/大乐透不同列）"""
        lt = self._lottery_type.get()
        if lt == "ssq":
            cols = ("期号","开奖日期","红球1","红球2","红球3","红球4","红球5","红球6","蓝球","奖池金额","一等奖","二等奖","销售额","操作")
            col_widths = {"期号":70,"开奖日期":95,"红球1":55,"红球2":55,"红球3":55,
                          "红球4":55,"红球5":55,"红球6":55,"蓝球":55,
                          "奖池金额":120,"一等奖":80,"二等奖":80,"销售额":120,"操作":80}
            hdr_bg = "#CC2222"
            sel_bg = "#FFCCCC"
            even_bg = "#FFF5F5"
        else:
            cols = ("期号","开奖日期","前区1","前区2","前区3","前区4","前区5","后区1","后区2","奖池金额","一等奖","二等奖","销售额","操作")
            col_widths = {"期号":70,"开奖日期":95,"前区1":55,"前区2":55,"前区3":55,
                          "前区4":55,"前区5":55,"后区1":55,"后区2":55,
                          "奖池金额":120,"一等奖":80,"二等奖":80,"销售额":120,"操作":80}
            hdr_bg = "#1565C0"
            sel_bg = "#CCE0FF"
            even_bg = "#F5F8FF"

        # 清理旧 tree
        if hasattr(self, 'tree') and self.tree:
            try:
                self.tree.destroy()
            except Exception:
                pass
        if hasattr(self, '_tree_ysb') and self._tree_ysb:
            try:
                self._tree_ysb.destroy()
            except Exception:
                pass
        if hasattr(self, '_tree_xsb') and self._tree_xsb:
            try:
                self._tree_xsb.destroy()
            except Exception:
                pass

        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=BG_CARD,
                        fieldbackground=BG_CARD,
                        rowheight=26,
                        font=("Microsoft YaHei", 9))
        style.configure("Treeview.Heading",
                        background=hdr_bg,
                        foreground="white",
                        font=("Microsoft YaHei", 9, "bold"),
                        relief="flat")
        style.map("Treeview", background=[("selected", sel_bg)])
        style.map("Treeview.Heading", background=[("active","#AA1111")])

        for col in cols:
            self.tree.heading(col, text=col, anchor="center",
                              command=lambda c=col: self._sort(c))
            self.tree.column(col, width=col_widths.get(col, 80), anchor="center", minwidth=40)

        self.tree.tag_configure("odd",  background="#FFFFFF")
        self.tree.tag_configure("even", background=even_bg)

        self._tree_ysb = ttk.Scrollbar(frame, orient="vertical",   command=self.tree.yview)
        self._tree_xsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self._tree_ysb.set, xscrollcommand=self._tree_xsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self._tree_ysb.grid(row=0, column=1, sticky="ns")
        self._tree_xsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Button-3>", self._on_right_click)

        self._sort_col = "期号"
        self._sort_asc = False

    def _rebuild_table_columns(self):
        """切换彩票类型时重建表格列"""
        self._build_tree_in_frame(self._table_frame)
        # 更新过滤器中的蓝球标签
        lt = self._lottery_type.get()
        if hasattr(self, '_blue_filter_lbl'):
            self._blue_filter_lbl.configure(text="蓝球：" if lt == "ssq" else "后区：")

    def _build_statusbar(self):
        bar = tk.Frame(self, bg="#EEEEEE", pady=4)
        bar.pack(fill="x", side="bottom")
        self.status_var = tk.StringVar(value="就绪")
        tk.Label(bar, textvariable=self.status_var,
                 bg="#EEEEEE", fg="#555", font=("Microsoft YaHei", 9)).pack(side="left", padx=10)
        # 抓取进度条（默认隐藏）
        self.fetch_pb_var = tk.IntVar(value=0)
        self.fetch_pb = ttk.Progressbar(bar, variable=self.fetch_pb_var,
                                        maximum=100, length=160, mode='determinate')
        self.fetch_pb.pack(side="left", padx=(0, 10))
        self.fetch_pb.pack_forget()  # 初始隐藏
        self.count_var = tk.StringVar()
        tk.Label(bar, textvariable=self.count_var,
                 bg="#EEEEEE", fg="#333", font=("Microsoft YaHei", 9, "bold")).pack(side="right", padx=10)

    # ── 数据加载 ────────────────────────────────────────────
    def _load_data(self):
        lt = self._lottery_type.get()
        if lt == "ssq":
            self._load_data_ssq()
        else:
            self._load_data_dlt()

    def _load_data_ssq(self):
        sql = ("SELECT issue,draw_date,red1,red2,red3,red4,red5,red6,blue,"
               "jackpot,prize1_count,prize2_count,sales FROM ssq_history WHERE 1=1")
        params = []

        v = self.var_issue_from.get().strip()
        if v:
            sql += " AND issue >= ?"
            params.append(v.zfill(5))
        v = self.var_issue_to.get().strip()
        if v:
            sql += " AND issue <= ?"
            params.append(v.zfill(5))
        v = self.var_date_from.get().strip()
        if v:
            sql += " AND draw_date >= ?"
            params.append(v)
        v = self.var_date_to.get().strip()
        if v:
            sql += " AND draw_date <= ?"
            params.append(v)
        v = self.var_red.get().strip()
        if v:
            try:
                n = int(v)
                sql += " AND (red1=? OR red2=? OR red3=? OR red4=? OR red5=? OR red6=?)"
                params.extend([n]*6)
            except:
                pass
        v = self.var_blue.get().strip()
        if v:
            try:
                sql += " AND blue=?"
                params.append(int(v))
            except:
                pass

        order = "ASC" if self._sort_asc else "DESC"
        col_map = {"期号":"issue","开奖日期":"draw_date","红球1":"red1","红球2":"red2",
                   "红球3":"red3","红球4":"red4","红球5":"red5","红球6":"red6",
                   "蓝球":"blue","奖池金额":"jackpot","一等奖":"prize1_count",
                   "二等奖":"prize2_count","销售额":"sales"}
        sort_col = col_map.get(self._sort_col, "issue")
        sql += f" ORDER BY {sort_col} {order}"

        rows = query_db(sql, params)
        self.tree.delete(*self.tree.get_children())
        self._current_rows = rows
        for i, row in enumerate(rows):
            tag = "odd" if i % 2 == 0 else "even"
            display = list(row) + ["编辑 | 删除"]
            self.tree.insert("", "end", iid=str(i), values=display, tags=(tag,))

        total = count_records("ssq")
        self.count_var.set(f"双色球 | 当前显示：{len(rows)} 条  |  数据库总计：{total} 条")
        self.status_var.set("查询完成")

    def _load_data_dlt(self):
        sql = ("SELECT issue,draw_date,red1,red2,red3,red4,red5,blue1,blue2,"
               "jackpot,prize1_count,prize2_count,sales FROM dlt_history WHERE 1=1")
        params = []

        v = self.var_issue_from.get().strip()
        if v:
            sql += " AND issue >= ?"
            params.append(v.zfill(5))
        v = self.var_issue_to.get().strip()
        if v:
            sql += " AND issue <= ?"
            params.append(v.zfill(5))
        v = self.var_date_from.get().strip()
        if v:
            sql += " AND draw_date >= ?"
            params.append(v)
        v = self.var_date_to.get().strip()
        if v:
            sql += " AND draw_date <= ?"
            params.append(v)
        v = self.var_red.get().strip()
        if v:
            try:
                n = int(v)
                sql += " AND (red1=? OR red2=? OR red3=? OR red4=? OR red5=?)"
                params.extend([n]*5)
            except:
                pass
        v = self.var_blue.get().strip()
        if v:
            try:
                sql += " AND (blue1=? OR blue2=?)"
                params.extend([int(v)]*2)
            except:
                pass

        order = "ASC" if self._sort_asc else "DESC"
        col_map = {"期号":"issue","开奖日期":"draw_date",
                   "前区1":"red1","前区2":"red2","前区3":"red3","前区4":"red4","前区5":"red5",
                   "后区1":"blue1","后区2":"blue2",
                   "奖池金额":"jackpot","一等奖":"prize1_count","二等奖":"prize2_count","销售额":"sales"}
        sort_col = col_map.get(self._sort_col, "issue")
        sql += f" ORDER BY {sort_col} {order}"

        rows = query_db(sql, params)
        self.tree.delete(*self.tree.get_children())
        self._current_rows = rows
        for i, row in enumerate(rows):
            tag = "odd" if i % 2 == 0 else "even"
            display = list(row) + ["编辑 | 删除"]
            self.tree.insert("", "end", iid=str(i), values=display, tags=(tag,))

        total = count_records("dlt")
        self.count_var.set(f"大乐透 | 当前显示：{len(rows)} 条  |  数据库总计：{total} 条")
        self.status_var.set("查询完成")

    def _reset_filter(self):
        for v in (self.var_issue_from, self.var_issue_to,
                  self.var_date_from, self.var_date_to,
                  self.var_red, self.var_blue):
            v.set("")
        self._load_data()

    def _sort(self, col):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = False
        self._load_data()

    # ── 操作绑定 ────────────────────────────────────────────
    def _on_double_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            idx = int(item)
            self._edit_record(self._current_rows[idx])

    def _on_right_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
        self.tree.selection_set(item)
        idx = int(item)
        row = self._current_rows[idx]
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="✏️ 编辑", command=lambda: self._edit_record(row))
        menu.add_command(label="🗑️ 删除", command=lambda: self._delete_record(row))
        menu.post(event.x_root, event.y_root)

    # ── 手动录入 ────────────────────────────────────────────
    def _manual_add(self):
        self._open_form(None)

    def _edit_record(self, row):
        self._open_form(row)

    def _open_form(self, row):
        win = tk.Toplevel(self)
        win.title("录入 / 编辑记录")
        win.geometry("480x420")
        win.resizable(False, False)
        win.configure(bg=BG_MAIN)
        win.grab_set()

        fields = [
            ("期号*", "issue"),
            ("开奖日期* (YYYY-MM-DD)", "draw_date"),
            ("红球1*", "red1"), ("红球2*", "red2"), ("红球3*", "red3"),
            ("红球4*", "red4"), ("红球5*", "red5"), ("红球6*", "red6"),
            ("蓝球*", "blue"),
            ("奖池金额", "jackpot"),
            ("一等奖注数", "prize1_count"),
            ("二等奖注数", "prize2_count"),
            ("销售额", "sales"),
        ]
        col_names = ["issue","draw_date","red1","red2","red3","red4","red5","red6",
                     "blue","jackpot","prize1_count","prize2_count","sales"]
        vals = {}
        if row:
            for k, v in zip(col_names, row):
                vals[k] = v

        frame = tk.Frame(win, bg=BG_MAIN, padx=20, pady=10)
        frame.pack(fill="both", expand=True)

        entries = {}
        for r, (label, key) in enumerate(fields):
            tk.Label(frame, text=label, bg=BG_MAIN,
                     font=("Microsoft YaHei", 9), anchor="w", width=22).grid(
                row=r, column=0, sticky="w", pady=3)
            var = tk.StringVar(value=str(vals.get(key, "")))
            e = tk.Entry(frame, textvariable=var, font=("Microsoft YaHei", 9),
                         relief="solid", bd=1, width=20)
            e.grid(row=r, column=1, sticky="ew", pady=3, padx=(6,0))
            entries[key] = var
        frame.columnconfigure(1, weight=1)

        def save():
            try:
                issue = entries["issue"].get().strip().zfill(5)
                draw_date = entries["draw_date"].get().strip()
                reds = [int(entries[f"red{i}"].get()) for i in range(1,7)]
                blue = int(entries["blue"].get())
                jackpot = entries["jackpot"].get().strip()
                p1 = entries["prize1_count"].get().strip()
                p2 = entries["prize2_count"].get().strip()
                sales = entries["sales"].get().strip()
                if not issue or not draw_date:
                    messagebox.showerror("错误", "期号和日期为必填项", parent=win)
                    return
                upsert_record((issue, draw_date, *reds, blue,
                               jackpot, int(p1) if p1 else 0,
                               int(p2) if p2 else 0, sales))
                self._load_data()
                self.status_var.set(f"期号 {issue} 已保存")
                win.destroy()
            except Exception as ex:
                messagebox.showerror("错误", f"数据格式有误：{ex}", parent=win)

        btn_frame = tk.Frame(win, bg=BG_MAIN)
        btn_frame.pack(pady=8)
        tk.Button(btn_frame, text="💾 保存", command=save,
                  bg=ACCENT, fg="white", relief="flat",
                  font=("Microsoft YaHei", 9, "bold"),
                  padx=20, pady=5, cursor="hand2").pack(side="left", padx=8)
        tk.Button(btn_frame, text="✖ 取消", command=win.destroy,
                  bg="#888", fg="white", relief="flat",
                  font=("Microsoft YaHei", 10),
                  padx=20, pady=5, cursor="hand2").pack(side="left", padx=8)

    def _delete_record(self, row):
        issue = row[0]
        if messagebox.askyesno("确认删除", f"确定要删除期号 {issue} 的记录吗？"):
            exec_db("DELETE FROM ssq_history WHERE issue=?", (issue,))
            self._load_data()
            self.status_var.set(f"已删除期号 {issue}")

    # ── 导入导出 ────────────────────────────────────────────
    def _download_template(self):
        # 获取当前彩票类型
        lottery_type = self._lottery_type.get()
        type_name = "大乐透" if lottery_type == "dlt" else "双色球"
        default_file = f"{type_name}导入模板.xlsx"
        
        path = filedialog.asksaveasfilename(
            title="保存导入模板",
            defaultextension=".xlsx",
            filetypes=[("Excel文件","*.xlsx")],
            initialfile=default_file
        )
        if not path:
            return
        export_template(path, lottery_type)
        messagebox.showinfo("成功", f"模板已保存至：\n{path}")

    def _import(self):
        path = filedialog.askopenfilename(
            title="选择要导入的Excel文件",
            filetypes=[("Excel文件","*.xlsx *.xls"),("所有文件","*.*")]
        )
        if not path:
            return
        try:
            count, errors = import_excel(path)
            msg = f"导入成功：{count} 条记录"
            if errors:
                msg += f"\n失败 {len(errors)} 行：\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    msg += f"\n...（共 {len(errors)} 个错误）"
            messagebox.showinfo("导入完成", msg)
            self._load_data()
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def _export(self):
        rows = getattr(self, "_current_rows", [])
        if not rows:
            messagebox.showwarning("提示", "当前无数据可导出")
            return
        
        # 获取当前彩票类型
        lottery_type = self._lottery_type.get()
        type_name = "大乐透" if lottery_type == "dlt" else "双色球"
        default_file = f"{type_name}历史数据.xlsx"
        
        path = filedialog.asksaveasfilename(
            title="导出数据",
            defaultextension=".xlsx",
            filetypes=[("Excel文件","*.xlsx")],
            initialfile=default_file
        )
        if not path:
            return
        export_data(path, rows, lottery_type)
        messagebox.showinfo("成功", f"已导出 {len(rows)} 条记录至：\n{path}")

    # ── 网络抓取 ────────────────────────────────────────────
    def _fetch_all(self):
        if not messagebox.askyesno("确认",
            "将从500彩票网抓取全部历史数据（双色球 + 大乐透）\n可能需要十几秒到一分钟，是否继续？"):
            return
        self._start_fetch_both(full=True)

    def _fetch_recent(self):
        self._start_fetch_both(full=False)

    def _start_fetch_both(self, full=False):
        """同时抓取双色球和大乐透数据（带进度条）"""
        # 显示进度条（不确定模式，旋转表示正在进行）
        self.fetch_pb.configure(mode='indeterminate')
        self.fetch_pb.pack(side="left", padx=(0, 10))
        self.fetch_pb.start(12)  # 每12ms一步，旋转式进度
        self._set_btns_state("disabled")

        def _safe_status(msg):
            """线程安全地更新状态文字"""
            def _do():
                self.status_var.set(msg)
            self.after(0, _do)

        def run():
            try:
                fd = _ensure_fetch_data()
                fdl = _ensure_fetch_dlt()
                if full:
                    _safe_status("⬇ 全量抓取中 - 双色球...")
                    fd.fetch_all(DB_PATH, progress_cb=_safe_status)
                    _safe_status("⬇ 全量抓取中 - 大乐透...")
                    fdl.fetch_all(DB_PATH, progress_cb=_safe_status)
                else:
                    _safe_status("🔄 增量更新中 - 双色球...")
                    fd.fetch_recent(DB_PATH, progress_cb=_safe_status)
                    _safe_status("🔄 增量更新中 - 大乐透...")
                    fdl.fetch_recent(DB_PATH, progress_cb=_safe_status)
            except Exception as e:
                _safe_status(f"抓取出错: {e}")
            finally:
                self.after(0, self._on_fetch_done)

        threading.Thread(target=run, daemon=True).start()

    def _start_fetch(self, func):
        """兼容旧调用，仅抓取双色球（带进度条）"""
        self.fetch_pb.configure(mode='indeterminate')
        self.fetch_pb.pack(side="left", padx=(0, 10))
        self.fetch_pb.start(12)
        self._set_btns_state("disabled")

        def _safe_status(msg):
            def _do():
                self.status_var.set(msg)
            self.after(0, _do)

        def run():
            try:
                func(DB_PATH, progress_cb=_safe_status)
            except Exception as e:
                _safe_status(f"抓取出错: {e}")
            finally:
                self.after(0, self._on_fetch_done)

        threading.Thread(target=run, daemon=True).start()

    def _on_fetch_done(self):
        self.fetch_pb.stop()
        self.fetch_pb.pack_forget()
        self._set_btns_state("normal")
        self._load_data()
        # 抓取完成后，自动输出三种预测结果到文本文件
        self._generate_prediction_txt()
        # 增量更新完成后，自动刷新预测结果
        if hasattr(self, "_pred_win") and self._pred_win and self._pred_win.winfo_exists():
            self._refresh_prediction(self._pred_win)
        else:
            # 弹窗提示并显示新预测
            self.after(300, self._show_prediction)

    def _generate_prediction_txt(self):
        """同时生成双色球和大乐透五种预测结果的纯文本文件"""
        import warnings
        warnings.filterwarnings("ignore")

        pred_dir = os.path.join(APP_DIR, "预测结果")
        os.makedirs(pred_dir, exist_ok=True)
        W = 64

        methods_to_try = [
            ("linear",        "线性回归"),
            ("entropy",       "熵最大化"),
            ("lstm",          "MLP神经网络"),
            ("xgboost",       "XGBoost"),
            ("gru",           "GRU序列"),
            ("frequency",     "频率加权"),
            ("gap",           "遗漏值"),
            ("hot_cold",      "冷热号平衡"),
            ("same_period",   "同期历史"),
            ("consecutive",   "连号跨度"),
            ("odd_even",      "奇偶比"),
            ("random_forest", "随机森林"),
            ("bayesian",      "贝叶斯"),
            ("genetic",       "遗传算法"),
        ]

        def _nums_ssq(result):
            if not result:
                return "[X] 数据不足"
            if "error" in result:
                return "[X] " + result.get("error", "预测失败")[:40]
            reds = " ".join([f"{r:02d}" for r in result.get("red_balls", [])])
            blue = result.get("blue_ball", "")
            blue_str = f"{blue:02d}" if isinstance(blue, int) else str(blue)
            return f"{reds} + {blue_str}"

        def _nums_dlt(result):
            if not result:
                return "[X] 数据不足"
            if "error" in result:
                return "[X] " + result.get("error", "预测失败")[:40]
            reds = " ".join([f"{r:02d}" for r in result.get("red_balls", [])])
            blues = result.get("blue_balls", [])
            blue_str = " ".join([f"{b:02d}" for b in blues]) if blues else str(result.get("blue_ball",""))
            return f"{reds} | {blue_str}"

        saved_paths = []

        # ── 双色球预测 ──────────────────────────────────────────
        try:
            results_ssq = {}
            base_ssq = None
            predict_mod = get_predict_module()  # 延迟加载
            for method, name in methods_to_try:
                try:
                    r = predict_mod.predict_next_issue(DB_PATH, method=method)
                    results_ssq[method] = r
                    if r and "error" not in r and base_ssq is None:
                        base_ssq = r
                except Exception as e:
                    results_ssq[method] = {"error": str(e)}

            if base_ssq is None:
                for r in results_ssq.values():
                    if r and "error" not in r:
                        base_ssq = r
                        break

            next_issue_ssq = base_ssq.get("next_issue", "未知") if base_ssq else "未知"
            history_ssq    = base_ssq.get("history_used", "?")  if base_ssq else "?"

            lines = []
            lines.append("=" * W)
            lines.append(f"  双色球  第 {next_issue_ssq} 期  预测报告")
            lines.append(f"  基于历史 {history_ssq} 期数据  |  生成时间: {self._get_current_time()}")
            lines.append("=" * W)
            lines.append("")
            lines.append("-" * W)
            for _, name in methods_to_try:
                lines.append(name)
            lines.append("")
            for method, _ in methods_to_try:
                lines.append(_nums_ssq(results_ssq.get(method)))
            lines.append("")
            lines.append("=" * W)
            lines.append("  [!] 以上预测仅供娱乐参考，彩票开奖结果完全随机，请理性购彩。")
            lines.append("=" * W)

            txt_path = os.path.join(pred_dir, f"SSQ预测_{next_issue_ssq}.txt")
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
            saved_paths.append(f"SSQ预测_{next_issue_ssq}.txt")
        except Exception as e:
            print(f"双色球预测文本生成失败: {e}")

        # ── 大乐透预测 ──────────────────────────────────────────
        try:
            results_dlt = {}
            base_dlt = None
            predict_dlt_mod = get_predict_dlt_module()  # 延迟加载
            for method, name in methods_to_try:
                try:
                    r = predict_dlt_mod.predict_next_issue(DB_PATH, method=method)
                    results_dlt[method] = r
                    if r and "error" not in r and base_dlt is None:
                        base_dlt = r
                except Exception as e:
                    results_dlt[method] = {"error": str(e)}

            if base_dlt is None:
                for r in results_dlt.values():
                    if r and "error" not in r:
                        base_dlt = r
                        break

            next_issue_dlt = base_dlt.get("next_issue", "未知") if base_dlt else "未知"
            history_dlt    = base_dlt.get("history_used", "?")  if base_dlt else "?"

            if base_dlt is None:
                print("大乐透无历史数据，跳过预测报告保存")
            else:
                lines = []
                lines.append("=" * W)
                lines.append(f"  大乐透  第 {next_issue_dlt} 期  预测报告")
                lines.append(f"  基于历史 {history_dlt} 期数据  |  生成时间: {self._get_current_time()}")
                lines.append("=" * W)
                lines.append("")
                lines.append("-" * W)
                for _, name in methods_to_try:
                    lines.append(name)
                lines.append("")
                for method, _ in methods_to_try:
                    lines.append(_nums_dlt(results_dlt.get(method)))
                lines.append("")
                lines.append("=" * W)
                lines.append("  [!] 以上预测仅供娱乐参考，彩票开奖结果完全随机，请理性购彩。")
                lines.append("=" * W)

                txt_path = os.path.join(pred_dir, f"DLT预测_{next_issue_dlt}.txt")
                with open(txt_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(lines))
                saved_paths.append(f"DLT预测_{next_issue_dlt}.txt")
        except Exception as e:
            print(f"大乐透预测文本生成失败: {e}")

        if saved_paths:
            self.status_var.set(f"预测报告已保存: {', '.join(saved_paths)}")
        else:
            self.status_var.set("预测报告生成失败，请检查数据")

    def _get_current_time(self):
        """获取当前时间字符串"""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _output_current_prediction(self, win):
        """一键输出预测结果：弹出美观进度窗口，后台执行完整自动预测流程"""
        self._run_auto_predict_ui()

    def _run_auto_predict_ui(self):
        """弹出精美进度窗口，后台线程跑 auto_predict 全流程（拉数据 + 五种预测 + 保存）"""
        import threading
        import warnings

        # 获取当前彩票类型
        lt = self._lottery_type.get()
        is_dlt = (lt == "dlt")
        title_text = "🔮  大乐透  自动预测" if is_dlt else "🔮  双色球  自动预测"

        # ── 窗口 ────────────────────────────────────────────────
        dlg = tk.Toplevel(self)
        dlg.title(title_text)
        dlg.geometry("540x820")
        dlg.resizable(False, False)
        dlg.configure(bg="#12002E")
        dlg.grab_set()          # 模态
        dlg.focus_force()

        # 居中
        dlg.update_idletasks()
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        x = (sw - 540) // 2
        y = (sh - 820) // 2
        dlg.geometry(f"540x820+{x}+{y}")

        # ── 顶部标题 ─────────────────────────────────────────────
        hdr = tk.Frame(dlg, bg="#2D0A6E", pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text=title_text,
                 bg="#2D0A6E", fg="white",
                 font=("Microsoft YaHei", 14, "bold")).pack()
        tk.Label(hdr, text="拉取最新数据 · 五种方案 · 一键输出",
                 bg="#2D0A6E", fg="#AA88FF",
                 font=("Microsoft YaHei", 9)).pack()

        # ── 总进度条 ─────────────────────────────────────────────
        prog_frame = tk.Frame(dlg, bg="#12002E", pady=12)
        prog_frame.pack(fill="x", padx=24)

        tk.Label(prog_frame, text="总进度", bg="#12002E", fg="#CCBBFF",
                 font=("Microsoft YaHei", 9)).pack(anchor="w")

        style = ttk.Style()
        style.theme_use("default")
        style.configure("AutoPred.Horizontal.TProgressbar",
                        troughcolor="#2A1055",
                        background="#7B2FFF",
                        thickness=16)
        total_pb = ttk.Progressbar(prog_frame, style="AutoPred.Horizontal.TProgressbar",
                                   orient="horizontal", length=490, mode="determinate",
                                   maximum=16)         # 1拉数据 + 14预测 + 1保存
        total_pb.pack(fill="x", pady=(2, 0))

        pct_var = tk.StringVar(value="0%")
        tk.Label(prog_frame, textvariable=pct_var, bg="#12002E", fg="#AA88FF",
                 font=("Microsoft YaHei", 9)).pack(anchor="e")

        # ── 五种方案小进度行 ─────────────────────────────────────
        steps_frame = tk.Frame(dlg, bg="#12002E")
        steps_frame.pack(fill="x", padx=24)

        _METHODS = [
            ("linear",        "📈 线性回归",     "#C62828"),
            ("entropy",       "🎲 熵最大化",     "#6A1B9A"),
            ("lstm",          "🧠 MLP神经网络",  "#7B1A40"),
            ("xgboost",       "🌲 XGBoost",      "#1B5E20"),
            ("gru",           "🔁 GRU序列",      "#0D47A1"),
            ("frequency",     "📊 频率加权",     "#BF360C"),
            ("gap",           "📉 遗漏值",       "#4A148C"),
            ("hot_cold",      "🌡️ 冷热号平衡",   "#880E4F"),
            ("same_period",   "📅 同期历史",     "#01579B"),
            ("consecutive",   "🔢 连号跨度",     "#1A237E"),
            ("odd_even",      "⚖️ 奇偶比",       "#004D40"),
            ("random_forest", "🌳 随机森林",     "#33691E"),
            ("bayesian",      "🔵 贝叶斯",       "#006064"),
            ("genetic",       "🧬 遗传算法",     "#E65100"),
        ]

        step_styles = {}   # method -> (label_var, icon_label)
        for method, label, color in _METHODS:
            row = tk.Frame(steps_frame, bg="#1C0040", pady=5)
            row.pack(fill="x", pady=2)

            icon_lbl = tk.Label(row, text="⏳", bg="#1C0040", fg="#888888",
                                font=("Segoe UI Emoji", 11), width=2)
            icon_lbl.pack(side="left", padx=(10, 4))

            tk.Label(row, text=label, bg="#1C0040", fg="#CCBBFF",
                     font=("Microsoft YaHei", 9), width=13, anchor="w").pack(side="left")

            result_var = tk.StringVar(value="等待中...")
            result_lbl = tk.Label(row, textvariable=result_var,
                                  bg="#1C0040", fg="#666699",
                                  font=("Consolas", 9), anchor="w")
            result_lbl.pack(side="left", fill="x", expand=True, padx=(4, 10))

            step_styles[method] = (result_var, result_lbl, icon_lbl)

        # ── 日志区 ───────────────────────────────────────────────
        log_frame = tk.Frame(dlg, bg="#12002E", pady=6)
        log_frame.pack(fill="both", expand=True, padx=24, pady=(4, 0))

        tk.Label(log_frame, text="日志", bg="#12002E", fg="#555577",
                 font=("Microsoft YaHei", 8)).pack(anchor="w")

        log_box = tk.Text(log_frame, bg="#0A0020", fg="#556655",
                          font=("Consolas", 8), height=5,
                          relief="flat", state="disabled",
                          insertbackground="white", wrap="word")
        log_box.pack(fill="both", expand=True)

        # ── 底部状态 + 关闭按钮 ─────────────────────────────────
        bot = tk.Frame(dlg, bg="#12002E", pady=10)
        bot.pack(fill="x", padx=24)

        status_var = tk.StringVar(value="准备中...")
        tk.Label(bot, textvariable=status_var, bg="#12002E", fg="#FFCC00",
                 font=("Microsoft YaHei", 9)).pack(side="left")

        close_btn = tk.Button(bot, text="关闭",
                              bg="#333355", fg="#AAAACC",
                              activebackground="#444477", activeforeground="white",
                              relief="flat", font=("Microsoft YaHei", 9),
                              padx=16, pady=4, cursor="hand2",
                              command=dlg.destroy, state="disabled")
        close_btn.pack(side="right")

        # ── 辅助函数（在主线程中更新 UI）────────────────────────
        _step = [0]   # 可变容器，线程可写

        def _log(msg):
            def _do():
                log_box.configure(state="normal")
                log_box.insert("end", msg + "\n")
                log_box.see("end")
                log_box.configure(state="disabled")
            dlg.after(0, _do)

        def _advance(msg=""):
            def _do():
                _step[0] += 1
                total_pb["value"] = _step[0]
                pct = int(_step[0] / total_pb["maximum"] * 100)
                pct_var.set(f"{pct}%")
                if msg:
                    status_var.set(msg)
            dlg.after(0, _do)

        def _set_step(method, result):
            """更新单个方案的显示"""
            def _do():
                result_var, result_lbl, icon_lbl = step_styles[method]
                if result and "error" not in result:
                    reds = " ".join([f"{r:02d}" for r in result.get("red_balls", [])])
                    # 根据彩票类型选择分隔符
                    if is_dlt:
                        # 大乐透：显示前后区
                        blues = result.get("blue_balls", [])
                        if blues:
                            blue_str = " ".join([f"{b:02d}" for b in blues])
                        else:
                            blue = result.get("blue_ball", "")
                            blue_str = f"{blue:02d}" if isinstance(blue, int) else str(blue)
                        result_var.set(f"{reds}  |  {blue_str}")
                    else:
                        # 双色球：显示红球+蓝球
                        blue = result.get("blue_ball", "")
                        blue_str = f"{blue:02d}" if isinstance(blue, int) else str(blue)
                        result_var.set(f"{reds}  +  {blue_str}")
                    result_lbl.configure(fg="#44FF88")
                    icon_lbl.configure(text="✅", fg="#44FF88")
                else:
                    err = result.get("error", "失败") if result else "数据不足"
                    result_var.set(f"[X] {err[:30]}")
                    result_lbl.configure(fg="#FF4444")
                    icon_lbl.configure(text="❌", fg="#FF4444")
            dlg.after(0, _do)

        def _set_running(method):
            def _do():
                result_var, result_lbl, icon_lbl = step_styles[method]
                result_var.set("预测中...")
                result_lbl.configure(fg="#FFAA00")
                icon_lbl.configure(text="⚡", fg="#FFAA00")
            dlg.after(0, _do)

        def _finish(txt_path, report_name):
            def _do():
                ltype = "大乐透" if is_dlt else "双色球"
                status_var.set(f"✅ 完成！{ltype}预测报告已保存至 预测结果/")
                close_btn.configure(state="normal", bg="#1B5E20", fg="white",
                                    activebackground="#2E7D32")
                self.status_var.set(f"预测报告已保存: {report_name}")
            dlg.after(0, _do)

        def _error(msg):
            def _do():
                status_var.set(f"❌ 出错：{msg[:60]}")
                close_btn.configure(state="normal", bg="#B71C1C", fg="white",
                                    activebackground="#C62828")
            dlg.after(0, _do)

        # ── 后台线程 ──────────────────────────────────────────────
        def _worker():
            try:
                warnings.filterwarnings("ignore")

                # 1. 初始化 & 拉取数据（双色球 + 大乐透）
                dlg.after(0, lambda: status_var.set("【1/3】正在拉取最新历史数据..."))
                _log("初始化数据库...")
                fd = _ensure_fetch_data()
                fdl = _ensure_fetch_dlt()
                fd.init_db(DB_PATH)
                fdl.init_dlt_db(DB_PATH)

                _log("拉取双色球历史数据中...")
                def _progress_cb(msg):
                    _log(str(msg))
                count_ssq = fd.fetch_all(DB_PATH, progress_cb=_progress_cb)
                _log(f"[OK] 双色球更新完成，新增 {count_ssq} 条记录")

                _log("拉取大乐透历史数据中...")
                count_dlt = fdl.fetch_all(DB_PATH, progress_cb=_progress_cb)
                _log(f"[OK] 大乐透更新完成，新增 {count_dlt} 条记录")
                _advance("【2/3】正在执行五种预测方案...")

                # 2. 根据当前彩种执行预测
                results_ssq = {}
                results_dlt = {}
                # ✅ 使用延迟加载的预测模块（解决启动慢问题）
                predict_mod = get_predict_dlt_module() if is_dlt else get_predict_module()
                prefix = "[DLT]" if is_dlt else "[SSQ]"
                
                # ✅ 创建进度回调函数，让预测方法能实时更新UI
                def _progress_cb(msg):
                    """预测过程中的实时进度回调"""
                    dlg.after(0, lambda m=msg: (
                        _log(f"  {m}"),
                        status_var.set(f"【2/3】{label} {m}")
                        if hasattr(label, 'split') else None
                    ))
                
                for method, label, _ in _METHODS:
                    _set_running(method)
                    _log(f"  {prefix} -> {label}...")
                    try:
                        # ✅ 传递进度回调，实时更新UI状态
                        r = predict_mod.predict_next_issue(DB_PATH, method=method,
                                                           progress_cb=_progress_cb)
                        if is_dlt:
                            results_dlt[method] = r
                        else:
                            results_ssq[method] = r
                        if r and "error" not in r:
                            _log(f"     [OK]")
                        else:
                            err = r.get("error", "未知") if r else "返回空"
                            _log(f"     [FAIL] {err[:40]}")
                    except Exception as ex:
                        err_str = str(ex)
                        if is_dlt:
                            results_dlt[method] = {"error": err_str}
                        else:
                            results_ssq[method] = {"error": err_str}
                        _log(f"     [FAIL] {err_str[:40]}")
                    # 使用对应的结果调用 _set_step（现在它会根据 is_dlt 正确格式化）
                    result = results_dlt[method] if is_dlt else results_ssq[method]
                    _set_step(method, result)
                    _advance()

                # 3. 构建 & 保存文件（根据当前彩种）
                dlg.after(0, lambda: status_var.set("【3/3】正在保存预测报告..."))
                W = 64
                pred_dir = os.path.join(APP_DIR, "预测结果")
                os.makedirs(pred_dir, exist_ok=True)

                def _fmt_ssq(r):
                    if r and "error" not in r:
                        reds = " ".join([f"{n:02d}" for n in r.get("red_balls", [])])
                        blue = r.get("blue_ball", "")
                        blue_str = f"{blue:02d}" if isinstance(blue, int) else str(blue)
                        return f"{reds} + {blue_str}"
                    err = r.get("error", "预测失败") if r else "数据不足"
                    return f"[X] {err[:30]}"

                def _fmt_dlt(r):
                    if r and "error" not in r:
                        reds = " ".join([f"{n:02d}" for n in r.get("red_balls", [])])
                        blues = r.get("blue_balls", [])
                        blue_str = " ".join([f"{b:02d}" for b in blues]) if blues else str(r.get("blue_ball",""))
                        return f"{reds} | {blue_str}"
                    err = r.get("error", "预测失败") if r else "数据不足"
                    return f"[X] {err[:30]}"

                # 根据 is_dlt 保存对应彩种的报告
                if is_dlt:
                    # 大乐透报告
                    base_dlt = next((results_dlt[m] for m, _, __ in _METHODS if results_dlt.get(m) and "error" not in results_dlt[m]), None)
                    next_dlt   = base_dlt.get("next_issue",   "未知") if base_dlt else "未知"
                    history_dlt = base_dlt.get("history_used", "?") if base_dlt else "?"
                    lines_dlt = []
                    lines_dlt.append("=" * W)
                    lines_dlt.append(f"  大乐透  第 {next_dlt} 期  预测报告")
                    lines_dlt.append(f"  基于历史 {history_dlt} 期数据  |  生成时间: {self._get_current_time()}")
                    lines_dlt.append("=" * W)
                    lines_dlt.append("")
                    lines_dlt.append("-" * W)
                    for method, label, _ in _METHODS:
                        lines_dlt.append(label.split(" ", 1)[-1])
                    lines_dlt.append("")
                    for method, _, __ in _METHODS:
                        lines_dlt.append(_fmt_dlt(results_dlt.get(method)))
                    lines_dlt.extend(["", "=" * W,
                        "  [!] 以上预测仅供娱乐参考，彩票开奖结果完全随机，请理性购彩。",
                        "=" * W])
                    path_dlt = os.path.join(pred_dir, f"DLT预测_{next_dlt}.txt")
                    with open(path_dlt, "w", encoding="utf-8") as f:
                        f.write("\n".join(lines_dlt))
                    _log(f"[OK] 大乐透报告已保存: DLT预测_{next_dlt}.txt")
                    saved_path = path_dlt
                else:
                    # 双色球报告
                    base_ssq = next((results_ssq[m] for m, _, __ in _METHODS if results_ssq.get(m) and "error" not in results_ssq[m]), None)
                    next_ssq   = base_ssq.get("next_issue",   "未知") if base_ssq else "未知"
                    history_ssq = base_ssq.get("history_used", "?") if base_ssq else "?"
                    lines_ssq = []
                    lines_ssq.append("=" * W)
                    lines_ssq.append(f"  双色球  第 {next_ssq} 期  预测报告")
                    lines_ssq.append(f"  基于历史 {history_ssq} 期数据  |  生成时间: {self._get_current_time()}")
                    lines_ssq.append("=" * W)
                    lines_ssq.append("")
                    lines_ssq.append("-" * W)
                    for method, label, _ in _METHODS:
                        lines_ssq.append(label.split(" ", 1)[-1])
                    lines_ssq.append("")
                    for method, _, __ in _METHODS:
                        lines_ssq.append(_fmt_ssq(results_ssq.get(method)))
                    lines_ssq.extend(["", "=" * W,
                        "  [!] 以上预测仅供娱乐参考，彩票开奖结果完全随机，请理性购彩。",
                        "=" * W])
                    path_ssq = os.path.join(pred_dir, f"SSQ预测_{next_ssq}.txt")
                    with open(path_ssq, "w", encoding="utf-8") as f:
                        f.write("\n".join(lines_ssq))
                    _log(f"[OK] 双色球报告已保存: SSQ预测_{next_ssq}.txt")
                    saved_path = path_ssq

                _advance("完成")
                _finish(saved_path, os.path.basename(saved_path).replace(".txt", "").replace("SSQ预测_", "SSQ/").replace("DLT预测_", "DLT/"))

            except Exception as ex:
                import traceback
                _log(traceback.format_exc())
                _error(str(ex))

        threading.Thread(target=_worker, daemon=True).start()

    def _set_btns_state(self, state):
        pass  # 简化：不禁用按钮，线程安全

    def _show_prediction(self):
        """打开预测结果窗口（若已打开则刷新）"""
        if hasattr(self, "_pred_win") and self._pred_win and self._pred_win.winfo_exists():
            self._refresh_prediction(self._pred_win)
            self._pred_win.lift()
            return

        lt = self._lottery_type.get()
        is_dlt = (lt == "dlt")

        win = tk.Toplevel(self)
        win.title("🔮 大乐透预测" if is_dlt else "🔮 双色球预测")
        win.geometry("760x720")
        win.resizable(True, True)
        win.configure(bg="#1A0533")
        self._pred_win = win

        # ── 顶部标题栏 ──────────────────────────────────────────
        title_bar = tk.Frame(win, bg="#3D1F7A", pady=8)
        title_bar.pack(fill="x")
        title_text = "🌟  大乐透  智能预测" if is_dlt else "🔮  双色球  智能预测"
        tk.Label(title_bar, text=title_text,
                 bg="#3D1F7A", fg="white",
                 font=("Microsoft YaHei", 14, "bold")).pack(side="left", padx=20, pady=2)
        # 一键输出当前预测结果按钮
        tk.Button(title_bar, text="💾 一键输出预测结果",
                 command=lambda: self._output_current_prediction(win),
                 bg="#2E7D32", fg="white",
                 activebackground="#1B5E20", activeforeground="white",
                 relief="flat", font=("Microsoft YaHei", 9, "bold"),
                 padx=12, pady=4, cursor="hand2").pack(side="right", padx=8)

        # ── 方案选择区：两行布局 ────────────────────────────────
        tab_bar = tk.Frame(win, bg="#2A1055", pady=6)
        tab_bar.pack(fill="x")

        self._pred_method = tk.StringVar(value="linear")

        # 第一行：统计类方法（原有4个）
        row1 = tk.Frame(tab_bar, bg="#2A1055")
        row1.pack()
        _stat_btns = [
            ("📈 线性回归", "linear",    "#5533AA", "#3311AA"),
            ("🎲 频率加权", "frequency", "#5533AA", "#3311AA"),
            ("📊 熵最大化", "entropy",   "#5533AA", "#3311AA"),
            ("🕳️ 遗漏值",  "gap",       "#5533AA", "#3311AA"),
        ]
        for txt, val, bg, sel in _stat_btns:
            tk.Radiobutton(row1, text=txt, variable=self._pred_method,
                           value=val, bg=bg, fg="white", selectcolor=sel,
                           activebackground=sel, activeforeground="white",
                           font=("Microsoft YaHei", 9),
                           indicatoron=0, padx=12, pady=4, relief="flat",
                           cursor="hand2",
                           command=lambda: self._refresh_prediction(win)
                           ).pack(side="left", padx=4, pady=2)

        # 分割线
        tk.Frame(tab_bar, bg="#443377", height=1).pack(fill="x", padx=12, pady=2)

        # 第二行：新增统计类方法（4个）
        row1b = tk.Frame(tab_bar, bg="#2A1055")
        row1b.pack()
        tk.Label(row1b, text="进阶统计：", bg="#2A1055", fg="#AA88FF",
                 font=("Microsoft YaHei", 8)).pack(side="left", padx=(8, 2))
        _stat2_btns = [
            ("🔥 冷热号平衡", "hot_cold",    "#5533AA", "#3311AA"),
            ("📅 同期历史",   "same_period", "#5533AA", "#3311AA"),
            ("🔗 连号跨度",   "consecutive", "#5533AA", "#3311AA"),
            ("⚖️ 奇偶比",     "odd_even",    "#5533AA", "#3311AA"),
        ]
        for txt, val, bg, sel in _stat2_btns:
            tk.Radiobutton(row1b, text=txt, variable=self._pred_method,
                           value=val, bg=bg, fg="white", selectcolor=sel,
                           activebackground=sel, activeforeground="white",
                           font=("Microsoft YaHei", 9),
                           indicatoron=0, padx=12, pady=4, relief="flat",
                           cursor="hand2",
                           command=lambda: self._refresh_prediction(win)
                           ).pack(side="left", padx=4, pady=2)

        # 分割线
        tk.Frame(tab_bar, bg="#443377", height=1).pack(fill="x", padx=12, pady=2)

        # 第三行：AI/ML 模型（原有3个 + 新增3个）
        row2 = tk.Frame(tab_bar, bg="#2A1055")
        row2.pack()
        tk.Label(row2, text="AI模型：", bg="#2A1055", fg="#AA88FF",
                 font=("Microsoft YaHei", 8)).pack(side="left", padx=(8, 2))
        _ml_btns = [
            ("🧠 MLP神经网络",  "lstm",          "#7B1A40", "#5C1230"),
            ("🌲 XGBoost",      "xgboost",       "#1B5E20", "#0A3D0E"),
            ("🔁 GRU序列模型",  "gru",           "#0D47A1", "#082F6B"),
            ("🌳 随机森林",     "random_forest", "#4A6741", "#2E4028"),
            ("🎯 贝叶斯",       "bayesian",      "#7B5E00", "#5A4400"),
            ("🧬 遗传算法",     "genetic",       "#5C4070", "#3D2B50"),
        ]
        for txt, val, bg, sel in _ml_btns:
            tk.Radiobutton(row2, text=txt, variable=self._pred_method,
                           value=val, bg=bg, fg="white", selectcolor=sel,
                           activebackground=sel, activeforeground="white",
                           font=("Microsoft YaHei", 9, "bold"),
                           indicatoron=0, padx=10, pady=4, relief="flat",
                           cursor="hand2",
                           command=lambda: self._refresh_prediction(win)
                           ).pack(side="left", padx=3, pady=2)

        # 内容区（可滚动）
        canvas = tk.Canvas(win, bg="#1A0533", highlightthickness=0)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        content = tk.Frame(canvas, bg="#1A0533")
        canvas_win = canvas.create_window((0, 0), window=content, anchor="nw")

        def on_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(canvas_win, width=canvas.winfo_width())
        content.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_win, width=e.width))

        # 鼠标滚轮支持
        def on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # 保存引用供刷新使用
        win._content = content
        win._canvas = canvas

        self._fill_prediction_content(content, win)

    def _refresh_prediction(self, win):
        """刷新预测窗口内容"""
        if not win.winfo_exists():
            return
        for w in win._content.winfo_children():
            w.destroy()
        # 所有方法统一走异步（后台线程+加载提示）
        self._fill_prediction_content(win._content, win, is_async=True)
        win._canvas.yview_moveto(0)

    def _show_lstm_result(self, parent, win, result, loading_label):
        """显示 LSTM 预测结果（异步回调，保持向后兼容）"""
        # 清除加载提示（旧版传入的是 loading_label）
        if loading_label and hasattr(loading_label, 'winfo_exists') and loading_label.winfo_exists():
            loading_label.destroy()
        # 直接展示结果
        self._show_prediction_result(parent, win, result, loading_frame=None)

    def _fill_prediction_content(self, parent, win, is_async=False, cached_result=None):
        """填充预测内容
        
        Args:
            parent: 父容器
            win: 窗口对象
            is_async: 是否异步执行（已废弃，所有方法统一走后台线程）
            cached_result: 预计算的结果（直接展示，不再重新计算）
        """
        # ── 如果有缓存结果，直接展示 ──
        if cached_result is not None:
            self._show_prediction_result(parent, win, cached_result, loading_frame=None)
            return

        method = self._pred_method.get()
        lt = self._lottery_type.get()
        is_dlt = (lt == "dlt")
        
        # 选择对应的预测模块（使用延迟加载）
        pred_module = get_predict_dlt_module() if is_dlt else get_predict_module()

        # ── 所有预测方法统一使用后台线程 + 加载提示 ──
        method_labels = {
            "linear":        "📈 线性回归计算中...",
            "frequency":     "🎲 频率加权随机生成中...",
            "entropy":       "📊 熵最大化分析中...",
            "gap":           "📉 遗漏值分析中...",
            "hot_cold":      "🌡️ 冷热号平衡分析中...",
            "same_period":   "📅 同期历史统计中...",
            "consecutive":   "🔢 连号跨度分析中...",
            "odd_even":      "⚖️ 奇偶比分析中...",
            "random_forest": "🌳 随机森林模型计算中...\n（首次约5秒）",
            "bayesian":      "🔵 贝叶斯概率计算中...",
            "genetic":       "🧬 遗传算法进化中...\n（约3-5秒）",
            "lstm":          "🧠 MLP 深度学习预测中...\n请稍候（约20秒）",
            "xgboost":       "🌲 XGBoost 训练中...\n首次需约30秒，之后秒级响应",
            "gru":           "🔁 GRU 序列模型训练中...\n首次需约2分钟，之后秒级响应",
        }

        loading_text = method_labels.get(method, "⚡ 预测计算中...")
        loading_frame = tk.Frame(parent, bg="#1A0533")
        loading_frame.pack(pady=40, fill="both", expand=True)

        # 加载动画：圆点闪烁
        loading_label = tk.Label(loading_frame, text=loading_text,
                                 bg="#1A0533", fg="#FFAA00",
                                 font=("Microsoft YaHei", 12),
                                 wraplength=400, justify="center")
        loading_label.pack(pady=(20, 5))

        # 进度提示动画
        dots_var = tk.StringVar(value="")
        dots_label = tk.Label(loading_frame, textvariable=dots_var,
                              bg="#1A0533", fg="#888888",
                              font=("Microsoft YaHei", 10))
        dots_label.pack()

        def _animate_dots():
            current = dots_var.get()
            if loading_frame.winfo_exists():
                if len(current) >= 10:
                    dots_var.set("")
                else:
                    dots_var.set(current + " .")
                dots_label.after(300, _animate_dots)

        _animate_dots()
        
        def run_in_thread():
            try:
                result = pred_module.predict_next_issue(DB_PATH, method=method)
            except Exception as e:
                import traceback
                traceback.print_exc()
                result = {"error": f"{e}"}
            # 在主线程更新 UI
            parent.after(0, lambda: self._show_prediction_result(parent, win, result, loading_frame))
        
        import threading
        t = threading.Thread(target=run_in_thread, daemon=True)
        t.start()
        return

    def _show_prediction_result(self, parent, win, result, loading_frame):
        """显示预测结果（异步回调）"""
        # 清除加载提示
        if loading_frame and hasattr(loading_frame, 'winfo_exists') and loading_frame.winfo_exists():
            loading_frame.destroy()

        # 获取当前方法名和彩票类型
        method = self._pred_method.get()
        is_dlt = (self._lottery_type.get() == "dlt")

        if not result or "error" in result:
            error_msg = result.get("error", "数据不足，请先抓取历史数据") if result else "数据不足，请先抓取历史数据"
            tk.Label(parent, text=f"⚠️  预测失败: {error_msg}",
                     bg="#1A0533", fg="#FF6644",
                     font=("Microsoft YaHei", 11)).pack(pady=40)
            return

        pad = dict(padx=20, pady=6)

        # ── 期号提示 ──
        issue_frame = tk.Frame(parent, bg="#2A1055", pady=8, padx=16,
                               highlightbackground="#5533AA", highlightthickness=1)
        issue_frame.pack(fill="x", **pad)
        tk.Label(issue_frame, text=f"预测期号：第  {result['next_issue']}  期",
                 bg="#2A1055", fg="#CCAAFF",
                 font=("Microsoft YaHei", 12, "bold")).pack(side="left")
        tk.Label(issue_frame, text=f"（基于全部 {result['history_used']} 期数据）",
                 bg="#2A1055", fg="#8866CC",
                 font=("Microsoft YaHei", 9)).pack(side="left", padx=10)

        # ── 预测号码展示 ──
        balls_frame = tk.Frame(parent, bg="#1A0533")
        balls_frame.pack(pady=(10, 4))

        tk.Label(balls_frame, text="预 测 号 码",
                 bg="#1A0533", fg="#FFFFFF",
                 font=("Microsoft YaHei", 11, "bold")).pack()

        nums_row = tk.Frame(balls_frame, bg="#1A0533")
        nums_row.pack(pady=8)

        # 红球（双色球6个 / 大乐透5个前区）
        for num in result["red_balls"]:
            lbl = tk.Label(nums_row, text=f"{num:02d}",
                           bg=RED_BG, fg="white",
                           font=("Microsoft YaHei", 14, "bold"),
                           width=3, height=1, relief="flat",
                           highlightbackground="#CC0000", highlightthickness=2)
            lbl.pack(side="left", padx=5, pady=4)
            lbl.configure(padx=6, pady=6)

        # 加号分隔
        tk.Label(nums_row, text=" + ",
                 bg="#1A0533", fg="#888888",
                 font=("Microsoft YaHei", 14, "bold")).pack(side="left")

        # 蓝球：双色球1个 / 大乐透2个后区
        if is_dlt:
            for bnum in result.get("blue_balls", []):
                blue_lbl = tk.Label(nums_row, text=f"{bnum:02d}",
                                    bg=BLUE_BG, fg="white",
                                    font=("Microsoft YaHei", 14, "bold"),
                                    width=3, height=1, relief="flat",
                                    highlightbackground="#0033AA", highlightthickness=2)
                blue_lbl.configure(padx=6, pady=6)
                blue_lbl.pack(side="left", padx=5, pady=4)
        else:
            blue_lbl = tk.Label(nums_row, text=f"{result['blue_ball']:02d}",
                                 bg=BLUE_BG, fg="white",
                                 font=("Microsoft YaHei", 14, "bold"),
                                 width=3, height=1, relief="flat",
                                 highlightbackground="#0033AA", highlightthickness=2)
            blue_lbl.configure(padx=6, pady=6)
            blue_lbl.pack(side="left", padx=5, pady=4)

        # ── 算法说明 ──
        method_frame = tk.Frame(parent, bg="#221144", pady=6, padx=16,
                                highlightbackground="#443377", highlightthickness=1)
        method_frame.pack(fill="x", **pad)
        tk.Label(method_frame, text=f"📐  算法：{result['method']}",
                 bg="#221144", fg="#AABBFF",
                 font=("Microsoft YaHei", 9)).pack(anchor="w")

        # ── 各位置详情（仅线性回归显示）──
        if method == "linear" and "detail" in result and isinstance(result["detail"], dict):
            detail_frame = tk.LabelFrame(parent, text=" 各号码位置原始预测值 ",
                                         bg="#1A0533", fg="#8866CC",
                                         font=("Microsoft YaHei", 9, "bold"),
                                         padx=12, pady=8,
                                         highlightbackground="#443377", highlightthickness=1)
            detail_frame.pack(fill="x", **pad)

            inner = tk.Frame(detail_frame, bg="#1A0533")
            inner.pack()
            col = 0
            for label, val in result["detail"].items():
                if "原始预测" in label:
                    pos = label.replace("红球位置", "位置").replace("原始预测", "")
                    box = tk.Frame(inner, bg="#2A1055", padx=8, pady=4,
                                   highlightbackground="#443377", highlightthickness=1)
                    box.grid(row=0, column=col, padx=4, pady=2)
                    tk.Label(box, text=pos, bg="#2A1055", fg="#AAAACC",
                             font=("Microsoft YaHei", 8)).pack()
                    tk.Label(box, text=f"{val}", bg="#2A1055", fg="#FF8888",
                             font=("Microsoft YaHei", 9, "bold")).pack()
                    col += 1
            # 蓝球原始
            if "blue_raw" in result:
                box = tk.Frame(inner, bg="#112255", padx=8, pady=4,
                               highlightbackground="#334477", highlightthickness=1)
                box.grid(row=0, column=col, padx=4, pady=2)
                tk.Label(box, text="蓝球", bg="#112255", fg="#AAAACC",
                         font=("Microsoft YaHei", 8)).pack()
                tk.Label(box, text=f"{result['blue_raw']}", bg="#112255", fg="#8899FF",
                         font=("Microsoft YaHei", 9, "bold")).pack()

        # ── 信息熵分析（仅熵最大化显示）──
        if method == "entropy" and "entropy_data" in result:
            ent = result["entropy_data"]
            entropy_frame = tk.LabelFrame(parent, text=" 📊 信息熵分析 ",
                                          bg="#1A0533", fg="#44AA88",
                                          font=("Microsoft YaHei", 9, "bold"),
                                          padx=12, pady=8)
            entropy_frame.pack(fill="x", **pad)

            # 熵值显示
            ent_inner = tk.Frame(entropy_frame, bg="#1A0533")
            ent_inner.pack(fill="x")

            # 红球熵
            red_frame = tk.Frame(ent_inner, bg="#2A1055", padx=10, pady=6,
                                 highlightbackground="#443377", highlightthickness=1)
            red_frame.pack(side="left", fill="x", expand=True, padx=4)
            tk.Label(red_frame, text="红球信息熵", bg="#2A1055", fg="#AAAACC",
                     font=("Microsoft YaHei", 9)).pack()
            tk.Label(red_frame, text=f"H = {ent['red_entropy']}", bg="#2A1055", fg="#88FFAA",
                     font=("Microsoft YaHei", 12, "bold")).pack()
            tk.Label(red_frame, text=f"相对熵: {ent['red_relative_entropy']*100:.1f}% (最大: {ent['red_max_entropy']})",
                     bg="#2A1055", fg="#88AA88", font=("Microsoft YaHei", 8)).pack()

            # 蓝球熵
            blue_frame = tk.Frame(ent_inner, bg="#112255", padx=10, pady=6,
                                  highlightbackground="#334477", highlightthickness=1)
            blue_frame.pack(side="left", fill="x", expand=True, padx=4)
            tk.Label(blue_frame, text="蓝球信息熵", bg="#112255", fg="#AAAACC",
                     font=("Microsoft YaHei", 9)).pack()
            tk.Label(blue_frame, text=f"H = {ent['blue_entropy']}", bg="#112255", fg="#88AAFF",
                     font=("Microsoft YaHei", 12, "bold")).pack()
            tk.Label(blue_frame, text=f"相对熵: {ent['blue_relative_entropy']*100:.1f}% (最大: {ent['blue_max_entropy']})",
                     bg="#112255", fg="#6688AA", font=("Microsoft YaHei", 8)).pack()

            # 变异系数
            cv_frame = tk.Frame(entropy_frame, bg="#1A0533", pady=6)
            cv_frame.pack(fill="x")
            tk.Label(cv_frame, text=f"红球变异系数: {ent['red_cv']:.4f}  |  蓝球变异系数: {ent['blue_cv']:.4f}",
                     bg="#1A0533", fg="#AAAAAA", font=("Microsoft YaHei", 9)).pack()
            tk.Label(cv_frame, text="变异系数越小，分布越均匀",
                     bg="#1A0533", fg="#666666", font=("Microsoft YaHei", 8)).pack()

            # 冷门号
            cold_frame = tk.Frame(entropy_frame, bg="#1A0533", pady=6)
            cold_frame.pack(fill="x")
            tk.Label(cold_frame, text="🥶 最冷门号码（低于期望频率）", bg="#1A0533", fg="#88CCFF",
                     font=("Microsoft YaHei", 9)).pack(anchor="w")
            cold_nums = tk.Frame(cold_frame, bg="#1A0533")
            cold_nums.pack(pady=4)
            for num in ent["cold_reds"]:
                freq = ent["red_freq"][num]
                exp = ent["expected_red_freq"]
                lbl = tk.Label(cold_nums, text=f"{num:02d}",
                               bg="#003344", fg="#88DDFF" if freq < exp else "#AAAAAA",
                               font=("Microsoft YaHei", 10, "bold" if freq < exp else "normal"),
                               padx=6, pady=2)
                lbl.pack(side="left", padx=2)
                # 提示偏离程度
                deviation = (exp - freq) / exp * 100 if exp > 0 else 0
                if deviation > 10:
                    tk.Label(cold_nums, text=f"↓{deviation:.0f}%", bg="#1A0533", fg="#66AACC",
                             font=("Microsoft YaHei", 7)).pack(side="left", padx=(0, 6))

        # ── 重新预测按钮（频率加权方案用）──
        if method == "frequency":
            refresh_btn_frame = tk.Frame(parent, bg="#1A0533")
            refresh_btn_frame.pack(pady=4)
            tk.Button(refresh_btn_frame, text="🔄 重新生成随机预测",
                      command=lambda: self._refresh_prediction(win),
                      bg="#5533AA", fg="white", relief="flat",
                      font=("Microsoft YaHei", 9), padx=15, pady=5,
                      cursor="hand2").pack()

        # ── 遗漏值分析（仅遗漏值方案显示）──
        if method == "gap" and "gap_data" in result:
            gap = result["gap_data"]
            gap_frame = tk.LabelFrame(parent, text=" 📊 遗漏值分析（Gap Analysis）",
                                      bg="#1A0533", fg="#FFAA44",
                                      font=("Microsoft YaHei", 9, "bold"),
                                      padx=12, pady=8)
            gap_frame.pack(fill="x", **pad)

            # 最大遗漏值显示
            max_frame = tk.Frame(gap_frame, bg="#1A0533")
            max_frame.pack(fill="x", pady=4)

            # 红球最大遗漏
            red_max_frame = tk.Frame(max_frame, bg="#3A2000", padx=10, pady=6,
                                     highlightbackground="#AA6600", highlightthickness=1)
            red_max_frame.pack(side="left", fill="x", expand=True, padx=4)
            tk.Label(red_max_frame, text="🔴 红球最大遗漏", bg="#3A2000", fg="#FFAA66",
                     font=("Microsoft YaHei", 9)).pack()
            tk.Label(red_max_frame, text=f"{gap['max_gap_red']} 号", bg="#3A2000", fg="#FFCC88",
                     font=("Microsoft YaHei", 16, "bold")).pack()
            tk.Label(red_max_frame, text=f"已遗漏 {gap['max_gap_red_value']} 期",
                     bg="#3A2000", fg="#CC8855", font=("Microsoft YaHei", 9)).pack()

            # 蓝球最大遗漏
            blue_max_frame = tk.Frame(max_frame, bg="#001A3A", padx=10, pady=6,
                                      highlightbackground="#3366AA", highlightthickness=1)
            blue_max_frame.pack(side="left", fill="x", expand=True, padx=4)
            tk.Label(blue_max_frame, text="🔵 蓝球最大遗漏", bg="#001A3A", fg="#88AAFF",
                     font=("Microsoft YaHei", 9)).pack()
            tk.Label(blue_max_frame, text=f"{gap['max_gap_blue']} 号", bg="#001A3A", fg="#AACCFF",
                     font=("Microsoft YaHei", 16, "bold")).pack()
            tk.Label(blue_max_frame, text=f"已遗漏 {gap['max_gap_blue_value']} 期",
                     bg="#001A3A", fg="#5588CC", font=("Microsoft YaHei", 9)).pack()

            # 平均遗漏
            avg_frame = tk.Frame(gap_frame, bg="#1A0533", pady=6)
            avg_frame.pack(fill="x")
            tk.Label(avg_frame, text=f"平均遗漏：红球 {gap['avg_gap_red']:.1f} 期  |  蓝球 {gap['avg_gap_blue']:.1f} 期",
                     bg="#1A0533", fg="#AAAAAA", font=("Microsoft YaHei", 9)).pack()

            # 最大遗漏号码 Top 10
            top_gap_frame = tk.Frame(gap_frame, bg="#1A0533", pady=6)
            top_gap_frame.pack(fill="x")
            tk.Label(top_gap_frame, text="📉 遗漏值最大的红球 Top 10", bg="#1A0533", fg="#CCAA66",
                     font=("Microsoft YaHei", 9)).pack(anchor="w")

            top_gap_nums = tk.Frame(top_gap_frame, bg="#1A0533")
            top_gap_nums.pack(pady=4)
            for i, num in enumerate(result.get("max_gap_reds", [])[:10]):
                gap_val = gap["red_gap"][num]
                is_selected = num in result["red_balls"]
                box = tk.Frame(top_gap_nums, bg="#442200" if is_selected else "#2A1A00",
                               padx=6, pady=3, highlightbackground="#AA6600" if is_selected else "#553300",
                               highlightthickness=2 if is_selected else 1)
                box.grid(row=0, column=i, padx=3)
                fg_color = "#FFFF00" if is_selected else "#CCAA66"
                tk.Label(box, text=f"{num:02d}", bg="#442200" if is_selected else "#2A1A00",
                         fg=fg_color, font=("Microsoft YaHei", 10, "bold" if is_selected else "normal")).pack()
                tk.Label(box, text=f"{gap_val}期", bg="#442200" if is_selected else "#2A1A00",
                         fg="#AA8855", font=("Microsoft YaHei", 7)).pack()

        # ── 红球热门号 ──
        red_label_text = " 🔥 前区热门号 TOP 10（全部历史频率） " if is_dlt else " 🔥 红球热门号 TOP 10（全部历史频率） "
        hot_frame = tk.LabelFrame(parent, text=red_label_text,
                                  bg="#1A0533", fg="#FF6644",
                                  font=("Microsoft YaHei", 9, "bold"),
                                  padx=12, pady=8)
        hot_frame.pack(fill="x", **pad)

        hot_inner = tk.Frame(hot_frame, bg="#1A0533")
        hot_inner.pack()
        for i, num in enumerate(result["hot_reds"]):
            freq_val = result["red_freq"][num]
            box = tk.Frame(hot_inner, bg="#3A0000", padx=6, pady=3,
                           highlightbackground="#AA2200", highlightthickness=1)
            box.grid(row=0, column=i, padx=3)
            is_pred = num in result["red_balls"]
            fg_num = "#FFFF00" if is_pred else "#FFAAAA"
            tk.Label(box, text=f"{num:02d}", bg="#3A0000", fg=fg_num,
                     font=("Microsoft YaHei", 10, "bold" if is_pred else "normal")).pack()
            tk.Label(box, text=f"{freq_val}次", bg="#3A0000", fg="#AA6655",
                     font=("Microsoft YaHei", 7)).pack()

        # ── 蓝球热门号 ──
        blue_label_text = " 💙 后区热门号 TOP 5（全部历史频率） " if is_dlt else " 💙 蓝球热门号 TOP 5（全部历史频率） "
        blue_hot_frame = tk.LabelFrame(parent, text=blue_label_text,
                                       bg="#1A0533", fg="#6688FF",
                                       font=("Microsoft YaHei", 9, "bold"),
                                       padx=12, pady=8)
        blue_hot_frame.pack(fill="x", **pad)

        bh_inner = tk.Frame(blue_hot_frame, bg="#1A0533")
        bh_inner.pack()
        # 大乐透：blue_balls 为列表；双色球：blue_ball 为单值
        pred_blues = set(result.get("blue_balls", [])) if is_dlt else {result.get("blue_ball")}
        for i, num in enumerate(result["hot_blues"]):
            freq_val = result["blue_freq"][num]
            is_pred = (num in pred_blues)
            box = tk.Frame(bh_inner, bg="#001133", padx=10, pady=3,
                           highlightbackground="#224488", highlightthickness=1)
            box.grid(row=0, column=i, padx=6)
            fg_num = "#FFFF00" if is_pred else "#AABBFF"
            tk.Label(box, text=f"{num:02d}", bg="#001133", fg=fg_num,
                     font=("Microsoft YaHei", 11, "bold" if is_pred else "normal")).pack()
            tk.Label(box, text=f"{freq_val}次", bg="#001133", fg="#556699",
                     font=("Microsoft YaHei", 7)).pack()

        # ── 方案说明 ──
        explain_frame = tk.LabelFrame(parent, text=" 📖 预测方案说明 ",
                                      bg="#1A0533", fg="#88AA88",
                                      font=("Microsoft YaHei", 9, "bold"),
                                      padx=12, pady=8)
        explain_frame.pack(fill="x", **pad)

        explain_text = "正在使用此预测方案，结果仅供参考。"  # 兜底默认值
        if method == "linear":
            if is_dlt:
                explain_text = """线性回归方案（大乐透）：
• 对前区5个位置和后区2个位置分别建立时间序列
• 使用最小二乘法拟合趋势线 y = kx + b
• 外推预测下一期的理论落点
• 结果截断到合法范围（前区1-35，后区1-12）"""
            else:
                explain_text = """线性回归方案：
• 对红球6个位置和蓝球分别建立时间序列
• 使用最小二乘法拟合趋势线 y = kx + b
• 外推预测下一期的理论落点
• 结果截断到合法范围（红球1-33，蓝球1-16）"""
        elif method == "frequency":
            if is_dlt:
                explain_text = """频率加权方案（大乐透）：
• 统计所有号码的历史出现次数
• 出现次数越多，被选中的概率越高
• 前区采用不放回加权抽样（5个不重复）
• 后区采用不放回加权抽样（2个不重复）
• 每次点击「重新生成」会得到不同的随机结果"""
            else:
                explain_text = """频率加权方案：
• 统计所有号码的历史出现次数
• 出现次数越多，被选中的概率越高
• 红球采用不放回加权抽样（6个不重复）
• 蓝球采用加权随机抽样
• 每次点击「重新生成」会得到不同的随机结果"""
        elif method == "entropy":
            if is_dlt:
                explain_text = """熵最大化方案（大乐透）：
• 计算信息熵 H = -Σ(p_i × log₂p_i) 衡量分布随机性
• 熵越大表示分布越均匀（接近真随机）
• 选择历史出现频率最低的"冷门"前区/后区号码
• 前区选 5 个最冷门，后区选 2 个最冷门"""
            else:
                explain_text = """熵最大化方案：
• 计算信息熵 H = -Σ(p_i × log₂p_i) 衡量分布随机性
• 熵越大表示分布越均匀（接近真随机）
• 选择历史出现频率最低的"冷门"号码
• 目的是使整体分布趋向均匀，最大化熵值
• 科学意义：检验彩票是否真正随机"""
        elif method == "lstm":
            if is_dlt:
                explain_text = """ML 方案（大乐透）：
• 大乐透 ML 方案当前使用线性回归代替
• 前区号码：从 1-35 中选 5 个
• 后区号码：从 1-12 中选 2 个
• 后续版本将支持大乐透专用神经网络模型"""
            else:
                explain_text = """MLP 深度学习方案：
• 使用 MLP 神经网络学习号码出现规律
• 输入：全部历史数据的统计特征（109维）
  - 红/蓝球频率分布
  - 红/蓝球当前遗漏值
  - 红球奇偶/大小比例
  - 红球尾数分布
  - 最近30期热号
• 输出：预测下一期各号码出现的概率
• 红球：33维sigmoid，预测6个红球
• 蓝球：16维分类
• 基于全部 3431 期历史数据
• 模型保存在 models/ssq_mlp_model.pkl"""
        elif method == "hot_cold":
            if is_dlt:
                explain_text = """冷热号平衡方案（大乐透）：
• 近20期频繁出现（≥2次）的前区号为"热号"
• 近20期未出现（0次）的前区号为"冷号"
• 按 3热2冷 比例混合选取前区5个号码
• 后区按 1热1冷 比例选取2个号码
• 符合"热冷周期性交替"规律"""
            else:
                explain_text = """冷热号平衡方案：
• 近20期频繁出现（≥2次）的红球为"热号"
• 近20期未出现（0次）的红球为"冷号"
• 按 4热2冷 比例混合选取6个红球
• 蓝球从近期热门中随机选取
• 符合"热冷周期性交替"规律"""
        elif method == "same_period":
            if is_dlt:
                explain_text = """同期历史方案（大乐透）：
• 统计历史上与当前同月份开奖的号码
• 找出同月份中出现频率最高的前区/后区号码
• 基于"季节性周期规律"假设
• 参考期数越多，结果越稳定"""
            else:
                explain_text = """同期历史方案：
• 统计历史上与当前同月份开奖的号码
• 找出同月份中出现频率最高的红球和蓝球
• 基于"季节性周期规律"假设
• 若同月历史不足5期，则取近30期代替"""
        elif method == "consecutive":
            if is_dlt:
                explain_text = """连号跨度分析方案（大乐透）：
• 统计历史中连号对（相邻号码同时出现）的频次
• 以最常见连号对为"锚点"
• 计算历史平均跨度（最大号-最小号），按此范围补充其余号码
• 生成含合理连号结构的前区号码组合
• 后区号码取历史高频"""
            else:
                explain_text = """连号跨度分析方案：
• 统计历史中连号对（相邻号码同时出现）的频次
• 以最常见连号对为"锚点"
• 计算历史平均跨度（最大-最小），按此范围补充其余4个红球
• 生成含合理连号结构的号码组合"""
        elif method == "odd_even":
            if is_dlt:
                explain_text = """奇偶比/大小比方案（大乐透）：
• 统计历史最优奇偶比（奇数:偶数前区号）
• 统计历史最优大小比（≥18:≤17 前区号）
• 在满足约束的高频号中优先选取
• 后区号码取历史高频后区号"""
            else:
                explain_text = """奇偶比/大小比方案：
• 统计历史最优奇偶比（红球奇数:偶数）
• 统计历史最优大小比（≥17:≤16 红球）
• 双色球历史最优比例通常为 3:3 或 4:2
• 在满足约束的高频号中优先选取"""
        elif method == "random_forest":
            if is_dlt:
                explain_text = """随机森林方案（大乐透）：
• 使用 sklearn RandomForestClassifier
• 对每个前区号码（1-35）做二分类（出/不出）
• 输入特征：近20期频率、遗漏值、奇偶性、大小性
• 取预测概率最高的5个前区号码
• 后区号码取历史高频
• 需要训练样本 ≥50 期"""
            else:
                explain_text = """随机森林方案：
• 使用 sklearn RandomForestClassifier（100棵树）
• 对每个红球号码（1-33）做二分类（出/不出）
• 输入特征：近20期频率、遗漏值、奇偶性、大小性
• 取预测概率最高的6个红球
• 蓝球同样用随机森林预测
• 需要训练样本 ≥50 期"""
        elif method == "bayesian":
            if is_dlt:
                explain_text = """贝叶斯条件概率方案（大乐透）：
• 统计历史"上一期出现了X → 本期出现了Y"的条件共现频率
• 计算 P(本期出Y | 上期出X) 并加权求和
• 加入先验频率平滑，避免零概率问题
• 选取综合条件得分最高的5个前区号
• 后区号码取历史高频"""
            else:
                explain_text = """贝叶斯条件概率方案：
• 统计历史"上一期出现了X → 本期出现了Y"的条件共现频率
• 计算 P(本期出Y | 上期出X) 并加权求和
• 加入先验频率（历史频率×0.1）做平滑
• 选取综合条件得分最高的6个红球
• 蓝球依据上一期蓝球的条件共现频率预测"""
        elif method == "genetic":
            if is_dlt:
                explain_text = """遗传算法方案（大乐透）：
• 将前区号码组合视为"个体"（种群200个）
• 适应度 = 与近50期历史开奖的平均匹配前区号数
• 经过80代进化（选择→交叉→变异，变异率15%）
• 初始种群偏向历史高频号，加速收敛
• 最终取适应度最高的个体作为预测结果"""
            else:
                explain_text = """遗传算法方案：
• 将红球号码组合视为"个体"（种群200个）
• 适应度 = 与近50期历史开奖的平均匹配红球数
• 经过80代进化（选择→交叉→变异，变异率15%）
• 初始种群偏向历史高频号，加速收敛
• 最终取适应度最高的个体作为预测结果"""
        elif method == "gap":
            if is_dlt:
                explain_text = """遗漏值分析方案（大乐透）：
• 统计每个号码距离上次出现已间隔多少期
• 选择遗漏值最大（最久未出）的前区5个+后区2个
• 基于"长期未出的号码该出了"的假设
• ⚠️ 注意：这是经典的"赌徒谬误"
• 科学事实：每次开奖独立，历史不影响未来"""
            else:
                explain_text = """遗漏值分析方案：
• 统计每个号码距离上次出现已间隔多少期
• 选择当前遗漏值最大（最久未出）的号码
• 基于"长期未出的号码该出了"的假设
• ⚠️ 注意：这是经典的"赌徒谬误"
• 科学事实：每次开奖独立，历史不影响未来"""

        tk.Label(explain_frame, text=explain_text,
                 bg="#1A0533", fg="#CCCCCC",
                 font=("Microsoft YaHei", 9),
                 justify="left", wraplength=600).pack(anchor="w")

        # ── 免责声明 ──
        tk.Label(parent,
                 text="⚠️  本预测仅基于数学模型，彩票开奖具有完全随机性，预测结果仅供参考，请理性购彩。",
                 bg="#1A0533", fg="#664455",
                 font=("Microsoft YaHei", 8),
                 wraplength=600).pack(pady=(8, 16))


    # ── 购彩账本 ────────────────────────────────────────────
    def _show_ledger(self):
        """打开购彩账本窗口"""
        if hasattr(self, "_ledger_win") and self._ledger_win and self._ledger_win.winfo_exists():
            self._ledger_win.lift()
            self._ledger_refresh()
            return

        win = tk.Toplevel(self)
        win.title("📒 购彩账本")
        win.geometry("960x700")
        win.minsize(800, 560)
        win.configure(bg="#0D1F0D")
        self._ledger_win = win

        # ── 顶部标题栏 ──
        title_bar = tk.Frame(win, bg="#1A4A1A", pady=8)
        title_bar.pack(fill="x")
        self._ledger_title_lbl = tk.Label(title_bar,
                 text="📒  购彩账本",
                 bg="#1A4A1A", fg="white",
                 font=("Microsoft YaHei", 13, "bold"))
        self._ledger_title_lbl.pack(side="left", padx=16)

        btn_kw = dict(relief="flat", font=("Microsoft YaHei", 9),
                      padx=10, pady=4, cursor="hand2")
        
        # 新增按钮 - 直接打开，使用当前彩票类型
        tk.Button(title_bar, text="➕ 新增双色球",
                  command=lambda: self._ledger_add_batch("ssq"),
                  bg="#AA2222", fg="white",
                  activebackground="#881111", activeforeground="white",
                  **btn_kw).pack(side="right", padx=6)
        tk.Button(title_bar, text="➕ 新增大乐透",
                  command=lambda: self._ledger_add_batch("dlt"),
                  bg="#1565C0", fg="white",
                  activebackground="#0D2A6A", activeforeground="white",
                  **btn_kw).pack(side="right", padx=6)
        tk.Button(title_bar, text="🔍 核对开奖结果",
                  command=self._ledger_check_selected,
                  bg="#1A5A8A", fg="white",
                  activebackground="#0D3A6A", activeforeground="white",
                  **btn_kw).pack(side="right", padx=6)

        self._ledger_type = "ssq"  # 默认显示双色球

        # ── 统计卡片区 ──
        self._ledger_summary_frame = tk.Frame(win, bg="#0D1F0D")
        self._ledger_summary_frame.pack(fill="x", padx=12, pady=(8, 4))

        # ── Notebook：购买记录 + 周/月/年统计 ──
        nb_style = ttk.Style()
        nb_style.configure("Ledger.TNotebook",
                            background="#0D1F0D", tabmargins=[2, 2, 2, 0])
        nb_style.configure("Ledger.TNotebook.Tab",
                            background="#1A3A1A", foreground="#AADDAA",
                            font=("Microsoft YaHei", 9, "bold"),
                            padding=[12, 5])
        nb_style.map("Ledger.TNotebook.Tab",
                     background=[("selected", "#2D8A2D")],
                     foreground=[("selected", "#FFFFFF")])

        nb = ttk.Notebook(win, style="Ledger.TNotebook")
        nb.pack(fill="both", expand=True, padx=12, pady=4)

        # ── Tab1：购买记录 ──
        list_frame = tk.Frame(nb, bg="#0D1F0D")
        nb.add(list_frame, text="📋 购买记录")

        # 增加"类型"列，同时显示双色球和大乐透
        cols = ("ID", "类型", "购买日期", "目标期号", "注数", "花费(元)",
                "核对状态", "开奖期号", "中奖注数", "最佳奖级", "备注")
        self._ledger_tree = ttk.Treeview(list_frame, columns=cols,
                                         show="headings", selectmode="browse")

        style = ttk.Style()
        style.configure("Ledger.Treeview",
                        background="#1A2A1A", fieldbackground="#1A2A1A",
                        rowheight=28, font=("Microsoft YaHei", 9),
                        foreground="#CCFFCC")
        style.configure("Ledger.Treeview.Heading",
                        background="#2A5A2A", foreground="white",
                        font=("Microsoft YaHei", 9, "bold"), relief="flat")
        style.map("Ledger.Treeview",
                  background=[("selected", "#2A6A2A")],
                  foreground=[("selected", "#FFFFFF")])
        self._ledger_tree.configure(style="Ledger.Treeview")

        col_widths = {"ID": 40, "类型": 70, "购买日期": 90, "目标期号": 80, "注数": 50,
                      "花费(元)": 70, "核对状态": 80, "开奖期号": 80,
                      "中奖注数": 80, "最佳奖级": 80, "备注": 120}
        for col in cols:
            self._ledger_tree.heading(col, text=col, anchor="center")
            self._ledger_tree.column(col, width=col_widths.get(col, 80),
                                     anchor="center", minwidth=40)

        self._ledger_tree.tag_configure("win",      background="#1A3A1A", foreground="#AAFFAA")
        self._ledger_tree.tag_configure("win_dlt",  background="#1A2A3A", foreground="#AAAFFF")
        self._ledger_tree.tag_configure("nocheck",  background="#2A2A1A", foreground="#DDDDAA")
        self._ledger_tree.tag_configure("lose",     background="#1A1A1A", foreground="#888888")
        self._ledger_tree.tag_configure("lose_dlt", background="#1A1A2A", foreground="#888888")

        ysb = ttk.Scrollbar(list_frame, orient="vertical",   command=self._ledger_tree.yview)
        xsb = ttk.Scrollbar(list_frame, orient="horizontal", command=self._ledger_tree.xview)
        self._ledger_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self._ledger_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        list_frame.rowconfigure(0, weight=1)
        list_frame.columnconfigure(0, weight=1)

        self._ledger_tree.bind("<Double-1>", self._ledger_on_double_click)
        self._ledger_tree.bind("<Button-3>",  self._ledger_on_right_click)

        # ── Tab2：周统计 ──
        week_frame = tk.Frame(nb, bg="#0D1F0D")
        nb.add(week_frame, text="📅 按周统计")
        self._ledger_week_frame = week_frame

        # ── Tab3：月统计 ──
        month_frame = tk.Frame(nb, bg="#0D1F0D")
        nb.add(month_frame, text="🗓 按月统计")
        self._ledger_month_frame = month_frame

        # ── Tab4：年统计 ──
        year_frame = tk.Frame(nb, bg="#0D1F0D")
        nb.add(year_frame, text="📆 按年统计")
        self._ledger_year_frame = year_frame

        # 切换标签时刷新统计内容
        self._ledger_nb = nb
        nb.bind("<<NotebookTabChanged>>", self._ledger_on_tab_change)

        # ── 底部状态栏 ──
        bot = tk.Frame(win, bg="#0A1A0A", pady=4)
        bot.pack(fill="x", side="bottom")
        self._ledger_status = tk.StringVar(value="就绪")
        tk.Label(bot, textvariable=self._ledger_status,
                 bg="#0A1A0A", fg="#88CC88",
                 font=("Microsoft YaHei", 9)).pack(side="left", padx=12)

        self._ledger_refresh()

    def _ledger_refresh(self):
        """刷新账本列表和统计卡片"""
        # 统计卡片 - 分别统计双色球和大乐透
        summ_ssq = _ensure_ledger().get_summary(DB_PATH, lottery_type="ssq")
        summ_dlt = _ensure_ledger().get_summary(DB_PATH, lottery_type="dlt")
        summ_total = {
            "total_cost": summ_ssq["total_cost"] + summ_dlt["total_cost"],
            "total_batches": summ_ssq["total_batches"] + summ_dlt["total_batches"],
            "total_tickets": summ_ssq["total_tickets"] + summ_dlt["total_tickets"],
            "total_win_tickets": summ_ssq["total_win_tickets"] + summ_dlt["total_win_tickets"],
            "checked_batches": summ_ssq["checked_batches"] + summ_dlt["checked_batches"],
            "prize_counts": summ_ssq["prize_counts"] + summ_dlt["prize_counts"],
        }
        for w in self._ledger_summary_frame.winfo_children():
            w.destroy()

        cards = [
            ("💰 总花费", f"¥ {summ_total['total_cost']:.0f}", "#3A6A3A"),
            ("📋 购买批次", f"{summ_total['total_batches']} 次", "#3A5A6A"),
            ("📝 总注数", f"{summ_total['total_tickets']} 注", "#4A4A2A"),
            ("🎉 中奖注数", f"{summ_total['total_win_tickets']} 注", "#6A3A2A"),
            ("✅ 已核对", f"{summ_total['checked_batches']} 次", "#3A3A6A"),
            ("🎱 双色球", f"{summ_ssq['total_batches']}批/{summ_ssq['total_cost']:.0f}元", "#5A2A2A"),
            ("🌟 大乐透", f"{summ_dlt['total_batches']}批/{summ_dlt['total_cost']:.0f}元", "#2A2A5A"),
        ]
        for title, val, bg in cards:
            card = tk.Frame(self._ledger_summary_frame, bg=bg, padx=14, pady=8,
                            highlightbackground="#2A4A2A", highlightthickness=1)
            card.pack(side="left", padx=6, pady=2)
            tk.Label(card, text=title, bg=bg, fg="#AADDAA",
                     font=("Microsoft YaHei", 8)).pack()
            tk.Label(card, text=val, bg=bg, fg="#FFFFFF",
                     font=("Microsoft YaHei", 12, "bold")).pack()

        # 刷新列表 - 同时加载两种类型的数据
        self._ledger_tree.delete(*self._ledger_tree.get_children())
        
        _prize_names_ssq = {1:"一等奖",2:"二等奖",3:"三等奖",
                            4:"四等奖",5:"五等奖",6:"六等奖"}
        _prize_names_dlt = {1:"一等奖",2:"二等奖",3:"三等奖",
                            4:"四等奖",5:"五等奖",6:"六等奖",
                            7:"七等奖",8:"八等奖",9:"九等奖"}
        
        all_batches = []
        for ltype in ("ssq", "dlt"):
            batches = lg.get_all_batches(DB_PATH, lottery_type=ltype)
            all_batches.extend(batches)
        
        # 按购买日期排序（最新的在前）
        all_batches.sort(key=lambda x: x[2], reverse=True)
        self._ledger_batches = all_batches
        
        for row in all_batches:
            bid, ltype, buy_date, target_issue, bet_count, total_cost, is_extra, \
                checked, draw_issue, notes, created_at, win_count, best_level = row

            type_str = "🎱 双色球" if ltype == "ssq" else "🌟 大乐透"
            if is_extra and ltype == "dlt":
                type_str += " +追"
            status_str = "✅ 已核对" if checked else "⏳ 未核对"
            draw_str   = draw_issue if draw_issue else "—"
            win_str    = f"{win_count} 注中奖" if checked else "—"

            if checked and best_level and best_level > 0:
                _prize_names = _prize_names_dlt if ltype == "dlt" else _prize_names_ssq
                best_str = _prize_names.get(best_level, f"奖级{best_level}")
            elif checked:
                best_str = "未中奖"
            else:
                best_str = "—"

            # 根据类型和中奖状态选择标签
            if checked and win_count and win_count > 0:
                tag = "win_dlt" if ltype == "dlt" else "win"
            elif checked:
                tag = "lose_dlt" if ltype == "dlt" else "lose"
            else:
                tag = "nocheck"

            self._ledger_tree.insert("", "end", iid=str(bid),
                values=(bid, type_str, buy_date, target_issue, bet_count,
                        f"¥{total_cost:.0f}", status_str, draw_str,
                        win_str, best_str, notes or ""),
                tags=(tag,))

        self._ledger_status.set(
            f"共 {len(all_batches)} 条记录  |  "
            f"总花费 ¥{summ_total['total_cost']:.0f}  |  "
            f"中奖 {summ_total['total_win_tickets']} 注"
        )

    def _ledger_on_double_click(self, event):
        item = self._ledger_tree.identify_row(event.y)
        if item:
            self._ledger_view_detail(int(item))

    def _ledger_on_right_click(self, event):
        item = self._ledger_tree.identify_row(event.y)
        if not item:
            return
        self._ledger_tree.selection_set(item)
        bid = int(item)
        menu = tk.Menu(self._ledger_win, tearoff=0)
        menu.add_command(label="📋 查看详情",
                         command=lambda: self._ledger_view_detail(bid))
        menu.add_command(label="🔍 核对开奖结果",
                         command=lambda: self._ledger_do_check(bid))
        menu.add_separator()
        menu.add_command(label="🗑️ 删除此记录",
                         command=lambda: self._ledger_delete(bid))
        menu.post(event.x_root, event.y_root)

    def _ledger_add_batch(self, ltype="ssq"):
        """新增购买批次对话框（可滚动，防止界面遮挡）
        ltype: 'ssq' 双色球 或 'dlt' 大乐透
        """
        cur_type = ltype  # 直接使用传入的类型，不再切换
        win = tk.Toplevel(self._ledger_win)
        win.title("➕ 新增购买记录 - " + ("🎱 双色球" if cur_type == "ssq" else "🌟 大乐透"))
        # 缩小高度，主体内容放入可滚动区域
        win.geometry("640x680")
        win.minsize(620, 500)
        win.configure(bg="#0D1F0D")
        win.grab_set()

        BG = "#0D1F0D"
        FG = "#CCFFCC"
        CARD = "#1A3A1A" if cur_type == "ssq" else "#1A2A3A"
        lkw = dict(bg=BG, fg=FG, font=("Microsoft YaHei", 9))
        ekw = dict(font=("Microsoft YaHei", 9), relief="solid", bd=1,
                   bg="#1A2A1A", fg="#FFFFFF", insertbackground="white")

        # ── 外层：Canvas滚动区 + 底部按钮固定区 ─────────────────────
        outer = tk.Frame(win, bg=BG)
        outer.pack(fill="both", expand=True)

        # Canvas + Scrollbar（滚动内容区）
        canvas = tk.Canvas(outer, bg=BG, highlightthickness=0,
                          scrollregion=(0, 0, 600, 1200))
        scr_v = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scr_v.set)

        scr_v.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # 内容画布（放在canvas内，随内容滚动）
        content = tk.Frame(canvas, bg=BG, width=620)
        canvas.create_window((0, 0), window=content, anchor="nw", tags="content")
        content.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_config(e):
            # 内容超长时显示滚动条，不长时自动适配高度
            cw = content.winfo_reqwidth()
            ch = content.winfo_reqheight()
            win_h = win.winfo_height()
            bar_h = win_h - 60  # 减去顶部/底部区域
            if ch > bar_h:
                canvas.itemconfig("content", width=604)
                scr_v.pack(side="right", fill="y")
            else:
                canvas.itemconfig("content", width=620)
                canvas.yview_moveto(0)
            canvas.configure(scrollregion=canvas.bbox("all"))

        canvas.bind("<Configure>", _on_canvas_config)
        content.bind("<Configure>",
                     lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # 支持鼠标滚轮滚动
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        win.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # 顶部鼠标拖动滚动
        def _drag_scroll(event):
            try:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass
        canvas.bind("<MouseWheel>", _drag_scroll)

        # 头部
        hdr = tk.Frame(content, bg="#1A4A1A" if cur_type == "ssq" else "#1A3A5A", pady=6)
        hdr.pack(fill="x")
        type_title = "🎱 双色球" if cur_type == "ssq" else "🌟 大乐透"
        hdr_lbl = tk.Label(hdr,
                 text=f"填写购买号码（12注，每注2元）  [{type_title}]" +
                      ("  |  可勾选「追加投注」" if cur_type == "dlt" else ""),
                 bg="#1A4A1A" if cur_type == "ssq" else "#1A3A5A", fg="white",
                 font=("Microsoft YaHei", 11, "bold"))
        hdr_lbl.pack(padx=16)

        # 不再需要类型切换区，直接使用传入的类型
        add_type_var = tk.StringVar(value=cur_type)

        body = tk.Frame(content, bg=BG, padx=16, pady=10)

        # 期号 & 日期
        row0 = tk.Frame(body, bg=BG)
        row0.pack(fill="x", pady=(0, 6))
        tk.Label(row0, text="购买日期：", **lkw).pack(side="left")
        var_date = tk.StringVar(value=date.today().strftime("%Y-%m-%d"))

        if _HAS_TKCALENDAR:
            _today = date.today()
            date_picker = _TkDateEntry(
                row0,
                textvariable=var_date,
                date_pattern="yyyy-mm-dd",
                year=_today.year, month=_today.month, day=_today.day,
                width=11,
                font=("Microsoft YaHei", 9),
                background="#1A4A1A",
                foreground="white",
                headersbackground="#2A6A2A",
                headersforeground="white",
                selectbackground="#3A8A3A",
                selectforeground="white",
                normalbackground="#1A2A1A",
                normalforeground="#CCFFCC",
                weekendbackground="#1A2A1A",
                weekendforeground="#AAFFAA",
                othermonthforeground="#446644",
                othermonthbackground="#0D1F0D",
                othermonthweforeground="#446644",
                othermonthwebackground="#0D1F0D",
                bordercolor="#2A5A2A",
                relief="solid",
                bd=1,
            )
            date_picker.pack(side="left", padx=(0, 20))
        else:
            date_entry_widget = tk.Entry(row0, textvariable=var_date, width=12, **ekw)
            date_entry_widget.pack(side="left", padx=(0, 4))

            def _open_cal_popup():
                import calendar as _cal
                popup = tk.Toplevel(win)
                popup.title("选择日期")
                popup.resizable(False, False)
                popup.configure(bg="#0D1F0D")
                popup.grab_set()
                try:
                    d = datetime.strptime(var_date.get(), "%Y-%m-%d").date()
                except ValueError:
                    d = date.today()
                cur_year = [d.year]
                cur_month = [d.month]

                hdr = tk.Frame(popup, bg="#1A4A1A")
                hdr.pack(fill="x")
                lbl_ym = tk.Label(hdr, text="", bg="#1A4A1A", fg="white",
                                  font=("Microsoft YaHei", 9, "bold"))
                lbl_ym.pack(side="left", expand=True)

                grid_frame = tk.Frame(popup, bg="#0D1F0D")
                grid_frame.pack(padx=8, pady=4)

                def _render():
                    lbl_ym.config(text=f"{cur_year[0]} 年 {cur_month[0]:02d} 月")
                    for w in grid_frame.winfo_children():
                        w.destroy()
                    for ci, wd in enumerate(["一","二","三","四","五","六","日"]):
                        tk.Label(grid_frame, text=wd, width=3,
                                 bg="#1A4A1A", fg="#AAFFAA",
                                 font=("Microsoft YaHei", 8, "bold")).grid(row=0, column=ci, pady=2)
                    cal = _cal.monthcalendar(cur_year[0], cur_month[0])
                    for ri, week in enumerate(cal, 1):
                        for ci, day in enumerate(week):
                            if day == 0:
                                tk.Label(grid_frame, text="", width=3, bg="#0D1F0D").grid(row=ri, column=ci)
                            else:
                                is_today = (day == date.today().day and
                                            cur_month[0] == date.today().month and
                                            cur_year[0] == date.today().year)
                                is_sel = (day == d.day and
                                          cur_month[0] == d.month and
                                          cur_year[0] == d.year)
                                bg_ = "#3A8A3A" if is_sel else ("#2A4A2A" if is_today else "#1A2A1A")
                                fg_ = "white" if is_sel else ("#AAFFAA" if is_today else "#CCFFCC")
                                btn = tk.Button(grid_frame, text=str(day), width=3,
                                                bg=bg_, fg=fg_, relief="flat",
                                                font=("Microsoft YaHei", 8),
                                                cursor="hand2",
                                                activebackground="#3A8A3A", activeforeground="white")
                                btn.grid(row=ri, column=ci, padx=1, pady=1)
                                def _pick(dy=day):
                                    var_date.set(f"{cur_year[0]}-{cur_month[0]:02d}-{dy:02d}")
                                    popup.destroy()
                                btn.configure(command=_pick)

                def _prev():
                    if cur_month[0] == 1:
                        cur_month[0] = 12; cur_year[0] -= 1
                    else:
                        cur_month[0] -= 1
                    _render()

                def _next():
                    if cur_month[0] == 12:
                        cur_month[0] = 1; cur_year[0] += 1
                    else:
                        cur_month[0] += 1
                    _render()

                tk.Button(hdr, text="◀", command=_prev,
                          bg="#1A4A1A", fg="white", relief="flat",
                          font=("Microsoft YaHei", 10), cursor="hand2").pack(side="left", padx=4)
                tk.Button(hdr, text="▶", command=_next,
                          bg="#1A4A1A", fg="white", relief="flat",
                          font=("Microsoft YaHei", 10), cursor="hand2").pack(side="right", padx=4)
                _render()

            tk.Button(row0, text="📅", command=_open_cal_popup,
                      bg="#2A5A2A", fg="white", relief="flat",
                      font=("Microsoft YaHei", 10), cursor="hand2",
                      padx=4, pady=1).pack(side="left", padx=(0, 16))

        tk.Label(row0, text="目标期号：", **lkw).pack(side="left")
        var_issue = tk.StringVar()

        def _update_issue_placeholder(ltype=None):
            lt = ltype or add_type_var.get()
            try:
                tbl = "dlt_history" if lt == "dlt" else "ssq_history"
                conn = sqlite3.connect(DB_PATH)
                cur = conn.cursor()
                cur.execute(f"SELECT issue FROM {tbl} ORDER BY issue DESC LIMIT 1")
                r = cur.fetchone()
                conn.close()
                if r:
                    var_issue.set(str(int(r[0]) + 1).zfill(5))
            except Exception:
                pass
        _update_issue_placeholder()

        tk.Entry(row0, textvariable=var_issue, width=8, **ekw).pack(side="left")

        # ── 预测文件选择区 ──
        select_frame = tk.LabelFrame(body, text=" 从预测结果模块选取 ",
                                     bg=CARD, fg="#FFD700",
                                     font=("Microsoft YaHei", 9, "bold"),
                                     padx=10, pady=8,
                                     highlightbackground="#5A4A2A", highlightthickness=1)
        select_frame.pack(fill="x", pady=(8, 6))

        def _get_pred_files(ltype):
            pred_dir = os.path.join(APP_DIR, "预测结果")
            if not os.path.exists(pred_dir):
                return []
            files = []
            # 根据彩种过滤文件
            # SSQ: SSQ预测_*.txt
            # DLT: DLT预测_*.txt
            for f in os.listdir(pred_dir):
                if f.endswith(".txt"):
                    if ltype == "ssq":
                        # 只取 SSQ预测_ 开头的文件
                        if f.startswith("SSQ预测_"):
                            issue = f.replace("SSQ预测_", "").replace(".txt", "")
                        else:
                            continue
                    else:  # dlt
                        # 只取 DLT预测_ 开头的文件
                        if f.startswith("DLT预测_"):
                            issue = f.replace("DLT预测_", "").replace(".txt", "")
                        else:
                            continue
                    filepath = os.path.join(pred_dir, f)
                    mtime = os.path.getmtime(filepath)
                    mtime_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
                    files.append((issue, mtime_str, filepath))
            files.sort(key=lambda x: x[0], reverse=True)
            return files

        def _parse_pred_file(filepath, ltype):
            predictions = []
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                # 找出所有包含号码的行（用 + 或 | 分隔）
                for ln in lines:
                    ln = ln.strip()
                    if not ln:
                        continue
                    # 双色球格式: "05 09 14 19 24 29 + 08"
                    # 大乐透格式: "05 09 14 19 24 + 08 12" 或 "05 09 14 19 24 | 08 12"
                    sep = '+' if '+' in ln else ('|' if '|' in ln else None)
                    if sep:
                        parts = ln.split(sep)
                        if len(parts) == 2:
                            nums = [int(x) for x in parts[0].split() if x.isdigit()]
                            bls = [int(x) for x in parts[1].split() if x.isdigit()]
                            if ltype == "ssq":
                                if len(nums) == 6 and bls:
                                    predictions.append({"reds": nums[:6], "blues": bls[:1]})
                            else:
                                if len(nums) == 5 and len(bls) >= 2:
                                    predictions.append({"reds": nums[:5], "blues": bls[:2]})
                return predictions[:12]

            except Exception:
                return []

        select_row = tk.Frame(select_frame, bg=CARD)
        select_row.pack(fill="x")
        tk.Label(select_row, text="选择预测记录（期号）：", bg=CARD, fg="#CCFFCC",
                 font=("Microsoft YaHei", 9)).pack(side="left")

        current_predictions = []
        method_vars = []
        method_cbs = []
        pred_files = [_get_pred_files(add_type_var.get())]
        pred_paths = {}
        preview_var = tk.StringVar(value="请选择预测记录并点击'加载预览'")

        cb_container = tk.Frame(select_frame, bg=CARD)
        cb_container.pack(fill="x")
        tk.Label(select_frame, textvariable=preview_var,
                 bg=CARD, fg="#88CC88", font=("Microsoft YaHei", 8),
                 wraplength=520, justify="left").pack(anchor="w", pady=(6, 0))

        def _refresh_pred_combo(ltype):
            nonlocal pred_files
            pred_files = _get_pred_files(ltype)
            if pred_files:
                opts = [f"第 {iss} 期  ({mt})" for iss, mt, _ in pred_files]
                combo_pred['values'] = opts
                var_sel_pred.set(opts[0])
                combo_pred.config(state="readonly")
                pred_paths.clear()
                pred_paths.update({f"第 {iss} 期  ({mt})": fp for iss, mt, fp in pred_files})
            else:
                combo_pred['values'] = ["暂无预测记录"]
                var_sel_pred.set("暂无预测记录")
                combo_pred.config(state="disabled")
                pred_paths.clear()

        def _load_preview():
            nonlocal current_predictions
            sel = var_sel_pred.get()
            if not sel or sel == "暂无预测记录":
                preview_var.set("请先选择预测记录")
                return
            fp = pred_paths.get(sel)
            if not fp:
                preview_var.set("预测文件不存在")
                return
            preds = _parse_pred_file(fp, cur_type)
            if not preds:
                preview_var.set("无法解析预测文件内容")
                return
            current_predictions = preds
            for cb in method_cbs:
                cb.destroy()
            method_vars.clear()
            method_cbs.clear()
            cb_frame = cb_container
            _all_method_names = ["线性回归","熵最大化","MLP神经网络","XGBoost","GRU序列",
                                  "冷热号平衡","同期历史","连号跨度","奇偶比","随机森林","贝叶斯","遗传算法"]
            method_names = _all_method_names[:len(preds)]
            if cur_type == "ssq":
                labels = [f"{i+1}. {n}: " + " ".join(f"{r:02d}" for r in p["reds"]) +
                          f" + {p['blues'][0]:02d}"
                          for i, (n, p) in enumerate(zip(method_names, preds))]
            else:
                labels = [f"{i+1}. {n}: " + " ".join(f"{r:02d}" for r in p["reds"]) +
                          f" | {p['blues'][0]:02d} {p['blues'][1]:02d}"
                          for i, (n, p) in enumerate(zip(method_names, preds))]
            for i, lbl in enumerate(labels):
                var = tk.BooleanVar(value=(i < 12))   # 前12个默认勾选，后面默认不勾
                method_vars.append(var)
                cb = tk.Checkbutton(cb_frame, text=lbl, variable=var,
                                   bg=CARD, fg="#CCFFCC", font=("Microsoft YaHei", 8),
                                   selectcolor="#2A5A2A", activebackground=CARD,
                                   activeforeground="#CCFFCC", anchor="w")
                cb.grid(row=i // 2, column=i % 2, sticky="w", padx=(0, 15), pady=1)
                method_cbs.append(cb)
            preview_var.set(f"已加载 {len(preds)} 组预测，可勾选后点击'选取并填入'（最多12注）")
            for iss, _, _ in pred_files:
                if var_sel_pred.get().startswith(f"第 {iss} 期  ("):
                    var_issue.set(iss); break

        def _fill_selected():
            if not current_predictions:
                messagebox.showwarning("提示", "请先点击'加载预览'", parent=win)
                return
            sel_idx = [i for i, v in enumerate(method_vars) if v.get()]
            if not sel_idx:
                messagebox.showwarning("提示", "请至少勾选一组", parent=win)
                return
            body_content_fill(sel_idx[:12])
            messagebox.showinfo("成功", f"已填入 {min(len(sel_idx),12)} 注（勾选 {len(sel_idx)} 组）", parent=win)

        def _fill_all():
            """一键填入全部预测数据（最多12组）"""
            if not current_predictions:
                messagebox.showwarning("提示", "请先点击'加载预览'", parent=win)
                return
            # 全部选中
            all_idx = list(range(min(12, len(current_predictions))))
            for var in method_vars:
                var.set(True)
            body_content_fill(all_idx)
            messagebox.showinfo("成功", f"已填入全部 {len(all_idx)} 注预测数据", parent=win)

        var_sel_pred = tk.StringVar()
        combo_pred = ttk.Combobox(select_row, textvariable=var_sel_pred,
                                  width=30, font=("Microsoft YaHei", 9), state="readonly")
        combo_pred.pack(side="left", padx=(5, 8))
        tk.Button(select_row, text="加载预览", command=_load_preview,
                  bg="#4A7C4A", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9), padx=10, pady=2, cursor="hand2").pack(side="left", padx=(0, 6))
        tk.Button(select_row, text="选取并填入", command=_fill_selected,
                  bg="#FF6B35", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9), padx=10, pady=2, cursor="hand2").pack(side="left", padx=(0, 6))
        tk.Button(select_row, text="填入全部注", command=_fill_all,
                  bg="#2E7D32", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9), padx=10, pady=2, cursor="hand2").pack(side="left", padx=(0, 6))
        tk.Button(select_row, text="🔄", command=lambda: _refresh_pred_combo(cur_type),
                  bg="#4A7C4A", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9), padx=8, pady=2, cursor="hand2").pack(side="left")
        tk.Label(select_frame, textvariable=preview_var,
                 bg=CARD, fg="#88CC88", font=("Microsoft YaHei", 8),
                 wraplength=520, justify="left").pack(anchor="w", pady=(6, 0))

        _refresh_pred_combo(cur_type)

        # ── 滚动区域：号码输入区（12注，滚动显示） ──
        ticket_vars = []

        # 外层固定高度容器（固定 620px，确保12个格子完整显示）
        tickets_outer = tk.Frame(body, bg=BG, height=620)
        tickets_outer.pack(fill="x", pady=(8, 0))
        tickets_outer.pack_propagate(False)   # 防止被内容撑大

        tickets_canvas = tk.Canvas(tickets_outer, bg=BG, highlightthickness=0,
                                   height=620)
        tickets_scroll = ttk.Scrollbar(tickets_outer, orient="vertical",
                                       command=tickets_canvas.yview)
        tickets_canvas.configure(yscrollcommand=tickets_scroll.set)
        tickets_scroll.pack(side="right", fill="y")
        tickets_canvas.pack(side="left", fill="both", expand=True)

        tickets_inner = tk.Frame(tickets_canvas, bg=BG)
        tickets_canvas_window = tickets_canvas.create_window(
            (0, 0), window=tickets_inner, anchor="nw")

        def _update_scroll(e):
            ri = tickets_canvas.bbox("all")
            if ri:
                tickets_canvas.configure(scrollregion=ri)
            tickets_canvas.itemconfig(tickets_canvas_window, width=tickets_canvas.winfo_width())

        tickets_inner.bind("<Configure>", _update_scroll)
        tickets_canvas.bind("<Configure>", _update_scroll)

        def body_content_fill(selected_idx):
            """填入选中的预测数据
            selected_idx: 勾选的方法索引列表，如 [0,1,3] 表示选了第0/1/3组方法
            逻辑：勾选的第1个方法 → 注单第1注，第2个方法 → 注单第2注，以此类推
            """
            for slot, pred_idx in enumerate(selected_idx):
                if slot >= len(ticket_vars):
                    break
                if pred_idx >= len(current_predictions):
                    continue
                pred = current_predictions[pred_idx]
                if cur_type == "ssq":
                    for j, v in enumerate(pred["reds"][:6], 1):
                        ticket_vars[slot][f"red{j}"].set(f"{v:02d}")
                    ticket_vars[slot]["blue"].set(f"{pred['blues'][0]:02d}")
                else:
                    for j, v in enumerate(pred["reds"][:5], 1):
                        ticket_vars[slot][f"red{j}"].set(f"{v:02d}")
                    ticket_vars[slot]["blue"].set(f"{pred['blues'][0]:02d}")
                    ticket_vars[slot]["blue2"].set(f"{pred['blues'][1]:02d}")

        def _render_tickets(ltype):
            for w in tickets_inner.winfo_children():
                w.destroy()
            ticket_vars.clear()
            for i in range(12):
                if ltype == "dlt":
                    frame_bg = "#1A2A1A"
                    fg_col = "#88FF88"
                    frame = tk.LabelFrame(tickets_inner, text=f"  第 {i+1} 注（后区2个号）  ",
                                  bg=frame_bg, fg=fg_col,
                                  font=("Microsoft YaHei", 8, "bold"),
                                  padx=6, pady=4,
                                  highlightbackground="#2A5A2A", highlightthickness=1)
                else:
                    frame_bg = CARD
                    fg_col = "#88FF88"
                    frame = tk.LabelFrame(tickets_inner, text=f"  第 {i+1} 注  ",
                                  bg=frame_bg, fg=fg_col,
                                  font=("Microsoft YaHei", 8, "bold"),
                                  padx=6, pady=4,
                                  highlightbackground="#2A5A2A", highlightthickness=1)
                frame.pack(fill="x", pady=1)
                inner = tk.Frame(frame, bg=frame_bg)
                inner.pack()
                tv = {}
                if ltype == "dlt":
                    tk.Label(inner, text="前区：", bg=frame_bg, fg="#FFAAAA",
                             font=("Microsoft YaHei", 7)).pack(side="left")
                    for j in range(1, 6):
                        v = tk.StringVar()
                        tk.Entry(inner, textvariable=v, width=3,
                                 font=("Microsoft YaHei", 9, "bold"),
                                 relief="solid", bd=1,
                                 bg="#3A1A1A", fg="#FF8888",
                                 insertbackground="white").pack(side="left", padx=2)
                        tv[f"red{j}"] = v
                    tk.Label(inner, text="后区：", bg=frame_bg, fg="#AAAAFF",
                             font=("Microsoft YaHei", 7)).pack(side="left", padx=(4, 0))
                    for j, lbl in enumerate(["后①", "后②"], 1):
                        v = tk.StringVar()
                        tk.Entry(inner, textvariable=v, width=3,
                                 font=("Microsoft YaHei", 9, "bold"),
                                 relief="solid", bd=1,
                                 bg="#1A1A3A", fg="#8888FF",
                                 insertbackground="white").pack(side="left", padx=2)
                        tv[f"blue{j}"] = v
                    tv["blue"] = tv.get("blue1")
                else:
                    tk.Label(inner, text="红球：", bg=CARD, fg="#FFAAAA",
                             font=("Microsoft YaHei", 7)).pack(side="left")
                    for j in range(1, 7):
                        v = tk.StringVar()
                        tk.Entry(inner, textvariable=v, width=3,
                                 font=("Microsoft YaHei", 9, "bold"),
                                 relief="solid", bd=1,
                                 bg="#3A1A1A", fg="#FF8888",
                                 insertbackground="white").pack(side="left", padx=2)
                        tv[f"red{j}"] = v
                    tk.Label(inner, text="蓝球：", bg=CARD, fg="#AAAAFF",
                             font=("Microsoft YaHei", 7)).pack(side="left", padx=(4, 0))
                    v = tk.StringVar()
                    tk.Entry(inner, textvariable=v, width=3,
                             font=("Microsoft YaHei", 9, "bold"),
                             relief="solid", bd=1,
                             bg="#1A1A3A", fg="#8888FF",
                             insertbackground="white").pack(side="left", padx=2)
                    tv["blue"] = v
                ticket_vars.append(tv)
            tickets_canvas.update_idletasks()
            ri = tickets_canvas.bbox("all")
            if ri:
                tickets_canvas.configure(scrollregion=ri)

        # 初始渲染（使用传入的类型）
        _render_tickets(cur_type)

        # 不再需要类型切换，删除 trace

        def _quick_fill():
            try:
                ltype = cur_type
                results = []
                if ltype == "dlt":
                    for m in ("linear", "entropy", "lstm", "xgboost", "gru"):
                        r = predict_dlt.predict_next_issue(DB_PATH, method=m)
                        if r and "error" not in r:
                            results.append(r)
                            if len(results) >= 12:
                                break
                else:
                    for m in ("linear", "entropy", "lstm", "xgboost", "gru"):
                        r = predict.predict_next_issue(DB_PATH, method=m)
                        if r and "error" not in r:
                            results.append(r)
                            if len(results) >= 12:
                                break
                if not results:
                    messagebox.showwarning("提示", "预测失败，请检查数据", parent=win)
                    return
                for i, res in enumerate(results[:12]):
                    if ltype == "dlt":
                        reds = res.get("red_balls", [])
                        blues = res.get("blue_balls", [])
                        for j, v in enumerate(reds[:5], 1):
                            ticket_vars[i][f"red{j}"].set(str(v))
                        ticket_vars[i]["blue"].set(str(blues[0]) if blues else "")
                        ticket_vars[i]["blue2"].set(str(blues[1]) if len(blues) > 1 else "")
                    else:
                        reds = res.get("red_balls", [])
                        blue = res.get("blue_ball", 0)
                        for j, v in enumerate(reds[:6], 1):
                            ticket_vars[i][f"red{j}"].set(str(v))
                        ticket_vars[i]["blue"].set(str(blue))
                var_issue.set(str(results[0].get("next_issue", var_issue.get())).zfill(5))
                messagebox.showinfo("成功", f"已填入 {len(results[:12])} 注预测号码", parent=win)
            except Exception as ex:
                messagebox.showwarning("提示", f"快速填入失败：{ex}", parent=win)

        tk.Button(body, text="⚡ 从当前期号预测结果快速填入12注",
                  command=_quick_fill,
                  bg="#5533AA", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9), padx=10, pady=3,
                  cursor="hand2").pack(anchor="w", pady=(4, 0))

        def _count_filled_tickets():
            """统计实际填写的注数（非空注）"""
            filled = 0
            for tv in ticket_vars:
                # 检查前区第一个球是否有值（以此判断该注是否已填写）
                if tv.get("red1") and tv["red1"].get().strip():
                    filled += 1
            return filled

        def _update_cost_label():
            n = _count_filled_tickets()
            extra = var_extra.get() if cur_type == "dlt" else False
            cost = n * 2 + (n if extra else 0)
            cost_label.config(
                text=f"预计花费：¥{cost}  ({n}注{' + 追加¥' + str(n) if extra else ''})"
            )

        def _on_extra_toggle():
            _update_cost_label()

        # 追加选项（仅大乐透）- 在 _on_extra_toggle 定义之后才能引用
        extra_row = tk.Frame(body, bg=BG)
        extra_row.pack(fill="x", pady=4)
        var_extra = tk.BooleanVar(value=False)
        extra_check = tk.Checkbutton(
            extra_row,
            text="☑ 追加投注（每注+1元，仅限大乐透）",
            variable=var_extra,
            bg=BG, fg="#FFD700",
            font=("Microsoft YaHei", 9),
            selectcolor="#1A2A1A",
            activebackground=BG, activeforeground="#FFD700",
            cursor="hand2" if cur_type == "dlt" else "",
            state="normal" if cur_type == "dlt" else "disabled",
            command=_on_extra_toggle
        )
        extra_check.pack(side="left")
        cost_label = tk.Label(extra_row, text="预计花费：¥24  (12注)",
                              bg=BG, fg="#88CC88",
                              font=("Microsoft YaHei", 9, "bold"))
        cost_label.pack(side="left", padx=(20, 0))

        # 备注
        note_row = tk.Frame(body, bg=BG)
        note_row.pack(fill="x", pady=4)
        tk.Label(note_row, text="备注：", **lkw).pack(side="left")
        var_notes = tk.StringVar()
        tk.Entry(note_row, textvariable=var_notes, width=40, **ekw).pack(side="left", padx=4)

        # 内容区加入滚动画布
        body.pack(fill="both", expand=True)

        def _save():
            tickets = []
            ltype = cur_type
            try:
                for i, tv in enumerate(ticket_vars):
                    # 大乐透：检查 red1 是否有值（空注则跳过）
                    if ltype == "dlt":
                        if not tv["red1"].get().strip():
                            continue
                        reds = []
                        for j in range(1, 6):
                            v = tv[f"red{j}"].get().strip()
                            if not v:
                                messagebox.showerror("错误",
                                    f"第{i+1}注前区{j}未填写", parent=win)
                                return
                            n = int(v)
                            if not (1 <= n <= 35):
                                messagebox.showerror("错误",
                                    f"第{i+1}注前区{j}必须在1~35之间", parent=win)
                                return
                            reds.append(n)
                        if len(set(reds)) != 5:
                            messagebox.showerror("错误",
                                f"第{i+1}注前区有重复数字", parent=win)
                            return
                        bv1 = tv["blue"].get().strip()
                        bv2 = tv["blue2"].get().strip()
                        if not bv1 or not bv2:
                            messagebox.showerror("错误",
                                f"第{i+1}注后区未填写完整（需填两个）", parent=win)
                            return
                        b1, b2 = int(bv1), int(bv2)
                        if not (1 <= b1 <= 12) or not (1 <= b2 <= 12):
                            messagebox.showerror("错误",
                                f"第{i+1}注后区必须在1~12之间", parent=win)
                            return
                        tickets.append({
                            "red1": reds[0], "red2": reds[1], "red3": reds[2],
                            "red4": reds[3], "red5": reds[4],
                            "blue": b1, "blue2": b2
                        })
                    else:
                        # 双色球：检查 red1 是否有值（空注则跳过）
                        if not tv["red1"].get().strip():
                            continue
                        reds = []
                        for j in range(1, 7):
                            v = tv[f"red{j}"].get().strip()
                            if not v:
                                messagebox.showerror("错误",
                                    f"第{i+1}注红球{j}未填写", parent=win)
                                return
                            n = int(v)
                            if not (1 <= n <= 33):
                                messagebox.showerror("错误",
                                    f"第{i+1}注红球{j}必须在1~33之间", parent=win)
                                return
                            reds.append(n)
                        if len(set(reds)) != 6:
                            messagebox.showerror("错误",
                                f"第{i+1}注红球有重复数字", parent=win)
                            return
                        bv = tv["blue"].get().strip()
                        if not bv:
                            messagebox.showerror("错误",
                                f"第{i+1}注蓝球未填写", parent=win)
                            return
                        blue = int(bv)
                        if not (1 <= blue <= 16):
                            messagebox.showerror("错误",
                                f"第{i+1}注蓝球必须在1~16之间", parent=win)
                            return
                        tickets.append({
                            "red1": reds[0], "red2": reds[1], "red3": reds[2],
                            "red4": reds[3], "red5": reds[4], "red6": reds[5],
                            "blue": blue
                        })

                if not tickets:
                    messagebox.showwarning("提示", "请至少填写1注再保存", parent=win)
                    return

                buy_date = var_date.get().strip()
                target_issue = var_issue.get().strip().zfill(5)
                if not buy_date or not target_issue:
                    messagebox.showerror("错误", "日期和期号不能为空", parent=win)
                    return

                try:
                    is_extra = var_extra.get()
                    # 大乐透：每注2元，追加每注+1元；双色球：每注2元
                    cost = len(tickets) * 2 + (len(tickets) if (is_extra and ltype == 'dlt') else 0)
                    _ensure_ledger().add_batch(DB_PATH, ltype, buy_date, target_issue,
                                     tickets, notes=var_notes.get().strip(),
                                     bet_count=len(tickets),
                                     is_extra=is_extra)
                    extra_str = "（含追加）" if is_extra and ltype == 'dlt' else ""
                    self._ledger_refresh()
                    self._ledger_status.set(f"已保存购买记录 [{'大乐透' if ltype == 'dlt' else '双色球'}]，期号 {target_issue}，花费 ¥{cost}{extra_str}")
                    win.destroy()
                except Exception as ex:
                    messagebox.showerror("保存失败", f"保存购买记录时出错：{ex}", parent=win)

            except ValueError as ex:
                messagebox.showerror("格式错误", f"请检查号码格式：{ex}", parent=win)
            except Exception as ex:
                messagebox.showerror("未知错误", f"发生错误：{ex}", parent=win)

        # ── 底部按钮区（固定在对话框底部，不随内容滚动）──────────────
        btn_outer = tk.Frame(win, bg="#0A1A0A", pady=6)
        btn_outer.pack(fill="x", side="bottom")
        tk.Button(btn_outer, text="💾 保存购买记录", command=_save,
                  bg="#2D8A2D", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9, "bold"),
                  padx=20, pady=5, cursor="hand2").pack(side="left", padx=10)
        tk.Button(btn_outer, text="✖ 取消", command=win.destroy,
                  bg="#555", fg="white", relief="flat",
                  font=("Microsoft YaHei", 10),
                  padx=20, pady=5, cursor="hand2").pack(side="left", padx=10)

    def _ledger_check_selected(self):
        """核对当前选中批次的开奖结果"""
        sel = self._ledger_tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先选择一条购买记录", parent=self._ledger_win)
            return
        bid = int(sel[0])
        self._ledger_do_check(bid)

    def _ledger_do_check(self, batch_id):
        """弹出核对窗口，填入或自动获取开奖结果"""
        conn = sqlite3.connect(DB_PATH)
        cur  = conn.cursor()
        cur.execute("SELECT target_issue, lottery_type FROM purchase_batch WHERE id=?",
                    (batch_id,))
        row = cur.fetchone()
        conn.close()
        if not row:
            return
        target_issue, ltype = row
        if ltype not in ("ssq", "dlt"):
            ltype = "ssq"

        win = tk.Toplevel(self._ledger_win)
        win.title("🔍 核对开奖结果")
        win.geometry("580x420")
        win.resizable(False, False)
        win.configure(bg="#0D1F2A")
        win.grab_set()

        BG = "#0D1F2A"
        lkw = dict(bg=BG, fg="#CCDDFF", font=("Microsoft YaHei", 9))
        ekw = dict(font=("Microsoft YaHei", 11, "bold"), relief="solid", bd=1,
                   bg="#1A2A3A", fg="#FFFFFF", insertbackground="white")

        hdr = tk.Frame(win, bg="#1A3A5A", pady=6)
        hdr.pack(fill="x")
        type_label = "🌟 大乐透" if ltype == "dlt" else "🎱 双色球"
        tk.Label(hdr, text=f"🔍 核对 [{type_label}] 期号 {target_issue} 的开奖结果",
                 bg="#1A3A5A", fg="white",
                 font=("Microsoft YaHei", 11, "bold")).pack(padx=16)

        body = tk.Frame(win, bg=BG, padx=20, pady=12)
        body.pack(fill="both", expand=True)

        var_draw_issue = tk.StringVar(value=target_issue)

        # SSQ: 6 red + 1 blue; DLT: 5 front + 2 back
        if ltype == "dlt":
            var_front = [tk.StringVar() for _ in range(5)]
            var_back  = [tk.StringVar() for _ in range(2)]
            var_blue_entries = []
        else:
            var_reds  = [tk.StringVar() for _ in range(6)]
            var_blue  = tk.StringVar()

        def _auto_fill():
            """从数据库自动填入目标期号的开奖号码"""
            try:
                conn2 = sqlite3.connect(DB_PATH)
                cur2  = conn2.cursor()
                if ltype == "dlt":
                    cur2.execute("""
                        SELECT issue, red1, red2, red3, red4, red5, blue1, blue2
                        FROM dlt_history WHERE issue=?""", (target_issue,))
                else:
                    cur2.execute("""
                        SELECT issue, red1, red2, red3, red4, red5, red6, blue
                        FROM ssq_history WHERE issue=?""", (target_issue,))
                r = cur2.fetchone()
                conn2.close()
                if r:
                    var_draw_issue.set(r[0])
                    if ltype == "dlt":
                        for i, v in enumerate(r[1:6]):
                            var_front[i].set(str(v))
                        for i, v in enumerate(r[6:8]):
                            var_back[i].set(str(v))
                    else:
                        for i, v in enumerate(r[1:7]):
                            var_reds[i].set(str(v))
                        var_blue.set(str(r[7]))
                    messagebox.showinfo("成功",
                        f"已自动填入第 {r[0]} 期开奖号码", parent=win)
                else:
                    messagebox.showwarning("未找到",
                        f"数据库中没有期号 {target_issue} 的开奖数据\n"
                        "请先更新历史数据，或手动填写", parent=win)
            except Exception as ex:
                messagebox.showerror("错误", str(ex), parent=win)

        tk.Button(body, text="⚡ 从数据库自动填入开奖号码",
                  command=_auto_fill,
                  bg="#1A5A8A", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9), padx=10, pady=3,
                  cursor="hand2").pack(anchor="w", pady=(0, 10))

        row_issue = tk.Frame(body, bg=BG)
        row_issue.pack(fill="x", pady=4)
        tk.Label(row_issue, text="开奖期号：", **lkw).pack(side="left")
        tk.Entry(row_issue, textvariable=var_draw_issue, width=10, **ekw).pack(side="left")

        if ltype == "dlt":
            row_front = tk.Frame(body, bg=BG)
            row_front.pack(fill="x", pady=6)
            tk.Label(row_front, text="开奖前区：", **lkw).pack(side="left")
            for i in range(5):
                tk.Entry(row_front, textvariable=var_front[i], width=4,
                         font=("Microsoft YaHei", 12, "bold"), relief="solid", bd=1,
                         bg="#3A1A1A", fg="#FF8888",
                         insertbackground="white").pack(side="left", padx=3)

            row_back = tk.Frame(body, bg=BG)
            row_back.pack(fill="x", pady=6)
            tk.Label(row_back, text="开奖后区：", **lkw).pack(side="left")
            for i in range(2):
                tk.Entry(row_back, textvariable=var_back[i], width=4,
                         font=("Microsoft YaHei", 12, "bold"), relief="solid", bd=1,
                         bg="#1A1A3A", fg="#8888FF",
                         insertbackground="white").pack(side="left", padx=3)
        else:
            row_red = tk.Frame(body, bg=BG)
            row_red.pack(fill="x", pady=6)
            tk.Label(row_red, text="开奖红球：", **lkw).pack(side="left")
            for i in range(6):
                tk.Entry(row_red, textvariable=var_reds[i], width=4,
                         font=("Microsoft YaHei", 12, "bold"), relief="solid", bd=1,
                         bg="#3A1A1A", fg="#FF8888",
                         insertbackground="white").pack(side="left", padx=3)

            row_blue = tk.Frame(body, bg=BG)
            row_blue.pack(fill="x", pady=6)
            tk.Label(row_blue, text="开奖蓝球：", **lkw).pack(side="left")
            tk.Entry(row_blue, textvariable=var_blue, width=4,
                     font=("Microsoft YaHei", 12, "bold"), relief="solid", bd=1,
                     bg="#1A1A3A", fg="#8888FF",
                     insertbackground="white").pack(side="left", padx=3)

        def _do_check():
            try:
                draw_issue = var_draw_issue.get().strip().zfill(5)
                if ltype == "dlt":
                    draw_front = [int(v.get()) for v in var_front]
                    draw_back  = [int(v.get()) for v in var_back]
                    draw_data = {"front": draw_front, "back": draw_back}
                else:
                    draw_reds = [int(v.get()) for v in var_reds]
                    draw_blue = int(var_blue.get())
                    draw_data = {"reds": draw_reds, "blue": draw_blue}

                results = _ensure_ledger().check_batch_result(
                    DB_PATH, batch_id, draw_issue, draw_data)
                self._ledger_refresh()
                win.destroy()
                self._ledger_view_detail(batch_id)

            except Exception as ex:
                messagebox.showerror("错误", f"核对失败：{ex}", parent=win)

        btn_f = tk.Frame(win, bg=BG)
        btn_f.pack(pady=8)
        tk.Button(btn_f, text="✅ 确认核对", command=_do_check,
                  bg="#1A5A8A", fg="white", relief="flat",
                  font=("Microsoft YaHei", 9, "bold"),
                  padx=20, pady=5, cursor="hand2").pack(side="left", padx=8)
        tk.Button(btn_f, text="✖ 取消", command=win.destroy,
                  bg="#555", fg="white", relief="flat",
                  font=("Microsoft YaHei", 10),
                  padx=20, pady=5, cursor="hand2").pack(side="left", padx=8)

        _auto_fill()

    def _ledger_view_detail(self, batch_id):
        """查看某批次详情（逐注展示中奖情况，支持 SSQ + DLT）"""
        tickets = _ensure_ledger().get_batch_tickets(DB_PATH, batch_id)

        conn = sqlite3.connect(DB_PATH)
        cur  = conn.cursor()
        cur.execute("""
            SELECT buy_date, target_issue, bet_count, checked, draw_issue,
                   notes, lottery_type, is_extra
            FROM purchase_batch WHERE id=?
        """, (batch_id,))
        batch = cur.fetchone()
        conn.close()
        if not batch:
            return

        buy_date, target_issue, bet_count, checked, draw_issue, notes, ltype, is_extra = batch
        if ltype not in ("ssq", "dlt"):
            ltype = "ssq"

        win = tk.Toplevel(self._ledger_win)
        win.title(f"📋 购买详情 — 期号 {target_issue}")
        win.geometry("720x620")
        win.resizable(True, True)
        win.configure(bg="#0D1F0D")

        type_str = "🌟 大乐透" if ltype == "dlt" else "🎱 双色球"
        if is_extra and ltype == "dlt":
            type_str += " +追"
        total_cost = bet_count * 2 + (bet_count if (is_extra and ltype == "dlt") else 0)
        hdr = tk.Frame(win, bg="#1A4A1A", pady=8)
        hdr.pack(fill="x")
        tk.Label(hdr,
                 text=f"📒 [{type_str}] 第 {target_issue} 期  |  购买日期 {buy_date}  |  花费 ¥{total_cost:.0f}",
                 bg="#1A4A1A", fg="white",
                 font=("Microsoft YaHei", 11, "bold")).pack(padx=16)

        # 若已核对，显示开奖号码
        draw_nums = None
        if checked and draw_issue:
            conn2 = sqlite3.connect(DB_PATH)
            cur2 = conn2.cursor()
            if ltype == "dlt":
                cur2.execute("""
                    SELECT red1,red2,red3,red4,red5,blue1,blue2
                    FROM dlt_history WHERE issue=?""", (draw_issue,))
            else:
                cur2.execute("""
                    SELECT red1,red2,red3,red4,red5,red6,blue
                    FROM ssq_history WHERE issue=?""", (draw_issue,))
            dr = cur2.fetchone()
            conn2.close()
            if dr:
                draw_nums = list(dr)

        canvas = tk.Canvas(win, bg="#0D1F0D", highlightthickness=0)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=12, pady=8)

        content = tk.Frame(canvas, bg="#0D1F0D")
        cw = canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(
            int(-1*(e.delta/120)), "units"))

        _prize_colors = {
            0: ("#1A1A1A", "#888888"),
            1: ("#5A0000", "#FFD700"),  # 一等奖
            2: ("#4A1A00", "#FFAA00"),  # 二等奖
            3: ("#3A2A00", "#FFFF00"),  # 三等奖
            4: ("#1A2A00", "#AAFFAA"),  # 四等奖
            5: ("#001A1A", "#88FFFF"),  # 五等奖
            6: ("#0A0A2A", "#AAAAFF"),  # 六等奖
            7: ("#2A0A2A", "#FFAAFF"),  # 七等奖
            8: ("#2A1A0A", "#FFEEAA"),  # 八等奖
            9: ("#1A2A0A", "#AAFFEE"),  # 九等奖
        }

        for t in tickets:
            (tno, r1, r2, r3, r4, r5, r6, blue, blue2,
             prize_level, prize_name, prize_ref,
             red_match, blue_match, front_hit, back_hit) = t

            lvl = prize_level if prize_level >= 0 else 0
            card_bg, card_fg = _prize_colors.get(lvl, ("#1A1A1A", "#CCCCCC"))

            card = tk.Frame(content, bg=card_bg, padx=12, pady=8,
                            highlightbackground=card_fg if lvl > 0 else "#333",
                            highlightthickness=2 if lvl > 0 else 1)
            card.pack(fill="x", pady=4, padx=4)

            row1 = tk.Frame(card, bg=card_bg)
            row1.pack(fill="x")
            tk.Label(row1, text=f"第 {tno} 注",
                     bg=card_bg, fg=card_fg,
                     font=("Microsoft YaHei", 9, "bold"),
                     width=6).pack(side="left")

            if ltype == "dlt":
                # DLT: 5 前区（红色）+ 2 后区（蓝色）
                front_nums = [r1, r2, r3, r4, r5]
                back_nums  = [blue, blue2] if blue2 else [blue]
                for n in front_nums:
                    hit = (draw_nums and n in draw_nums[:5])
                    bg_n = "#FF4444" if hit else "#880000"
                    tk.Label(row1, text=f"{n:02d}", bg=bg_n, fg="white",
                             font=("Microsoft YaHei", 9, "bold"),
                             padx=5, pady=2, relief="flat").pack(side="left", padx=2)
                tk.Label(row1, text=" | ", bg=card_bg, fg="#888",
                         font=("Microsoft YaHei", 10)).pack(side="left")
                for n in back_nums:
                    hit = (draw_nums and n in draw_nums[5:])
                    bg_b = "#2255FF" if hit else "#113388"
                    tk.Label(row1, text=f"{n:02d}", bg=bg_b, fg="white",
                             font=("Microsoft YaHei", 9, "bold"),
                             padx=5, pady=2, relief="flat").pack(side="left", padx=2)
            else:
                # SSQ: 6 红球 + 1 蓝球
                red_nums = [r1, r2, r3, r4, r5, r6]
                for n in red_nums:
                    hit = (draw_nums and n in draw_nums[:6])
                    bg_n = "#FF4444" if hit else "#880000"
                    tk.Label(row1, text=f"{n:02d}", bg=bg_n, fg="white",
                             font=("Microsoft YaHei", 9, "bold"),
                             padx=5, pady=2, relief="flat").pack(side="left", padx=2)
                tk.Label(row1, text=" + ", bg=card_bg, fg="#666",
                         font=("Microsoft YaHei", 10)).pack(side="left")
                hit = (draw_nums and blue == draw_nums[6])
                bg_b = "#2255FF" if hit else "#113388"
                tk.Label(row1, text=f"{blue:02d}", bg=bg_b, fg="white",
                         font=("Microsoft YaHei", 9, "bold"),
                         padx=5, pady=2, relief="flat").pack(side="left", padx=2)

            # 第二行：中奖信息
            if prize_level >= 0:
                row2 = tk.Frame(card, bg=card_bg)
                row2.pack(fill="x", pady=(4, 0))
                if prize_level == 0:
                    status_text = "❌ 未中奖"
                    status_color = "#666666"
                else:
                    status_text = f"🎉 {prize_name}  参考奖金：{prize_ref}"
                    status_color = card_fg
                if ltype == "dlt":
                    info = f"  前区 {front_hit} 个  后区 {back_hit} 个    {status_text}"
                else:
                    info = f"  红球 {red_match} 个  {'✔ 蓝球命中' if blue_match else '✗ 蓝球未中'}    {status_text}"
                tk.Label(row2, text=info, bg=card_bg, fg=status_color,
                         font=("Microsoft YaHei", 9)).pack(side="left", padx=6)

        # 底部总结
        if tickets:
            win_tickets = [t for t in tickets if t[9] > 0]
            summary_bg = "#1A3A1A" if win_tickets else "#1A1A1A"
            sumf = tk.Frame(content, bg=summary_bg, pady=8, padx=12,
                            highlightbackground="#2A5A2A", highlightthickness=1)
            sumf.pack(fill="x", pady=(8, 4), padx=4)
            if win_tickets:
                txt = f"🎉 本次购买共中奖 {len(win_tickets)} 注！  花费 ¥{total_cost:.0f}"
            else:
                if checked:
                    txt = f"😞 本次未中奖   花费 ¥{total_cost:.0f}"
                else:
                    txt = f"⏳ 尚未核对开奖结果   花费 ¥{total_cost:.0f}"
            tk.Label(sumf, text=txt, bg=summary_bg,
                     fg="#88FF88" if win_tickets else "#888888",
                     font=("Microsoft YaHei", 9, "bold")).pack()
            if notes:
                tk.Label(sumf, text=f"备注：{notes}", bg=summary_bg,
                         fg="#AAAAAA",
                         font=("Microsoft YaHei", 9)).pack()

    def _ledger_delete(self, batch_id):
        if messagebox.askyesno("确认删除", f"确定要删除 ID={batch_id} 的购买记录吗？",
                               parent=self._ledger_win):
            _ensure_ledger().delete_batch(DB_PATH, batch_id)
            self._ledger_refresh()
            self._ledger_status.set(f"已删除购买记录 ID={batch_id}")

    # ── 周/月/年统计 ─────────────────────────────────────────

    def _ledger_on_tab_change(self, event):
        """切换统计标签时刷新对应内容"""
        idx = self._ledger_nb.index("current")
        if idx == 1:
            self._ledger_render_period_stats(
                self._ledger_week_frame,
                _ensure_ledger().get_weekly_stats(DB_PATH, lottery_type=self._ledger_type),
                ("week_label", "week_start", "week_end", "cost", "batches", "win_tickets"),
                ("周次", "开始日期", "结束日期", "花费(元)", "购买次数", "中奖注数"),
                (100, 95, 95, 80, 80, 80),
                "周",
            )
        elif idx == 2:
            self._ledger_render_period_stats(
                self._ledger_month_frame,
                _ensure_ledger().get_monthly_stats(DB_PATH, lottery_type=self._ledger_type),
                ("month_label", "cost", "batches", "win_tickets"),
                ("月份", "花费(元)", "购买次数", "中奖注数"),
                (100, 100, 100, 100),
                "月",
            )
        elif idx == 3:
            self._ledger_render_period_stats(
                self._ledger_year_frame,
                _ensure_ledger().get_yearly_stats(DB_PATH, lottery_type=self._ledger_type),
                ("year_label", "cost", "batches", "win_tickets"),
                ("年份", "花费(元)", "购买次数", "中奖注数"),
                (120, 120, 120, 120),
                "年",
            )

    def _ledger_render_period_stats(self, parent, data, keys, headers, widths, period_name):
        """通用：渲染某个时间周期的统计表格 + 汇总卡片"""
        # 清空
        for w in parent.winfo_children():
            w.destroy()

        BG = "#0D1F0D"

        if not data:
            tk.Label(parent, text=f"暂无{period_name}统计数据，请先添加购买记录",
                     bg=BG, fg="#888888",
                     font=("Microsoft YaHei", 11)).pack(expand=True)
            return

        # ── 汇总卡片 ──────────────────────────────────────────
        total_cost    = sum(row["cost"]        for row in data)
        total_batches = sum(row["batches"]     for row in data)
        total_wins    = sum(row["win_tickets"] for row in data)
        period_count  = len(data)

        summary_frame = tk.Frame(parent, bg=BG)
        summary_frame.pack(fill="x", padx=12, pady=(8, 4))

        card_info = [
            ("📊 统计周期数",  f"{period_count} 个{period_name}", "#3A5A3A"),
            ("💰 合计花费",    f"¥ {total_cost:.0f}",             "#3A4A6A"),
            ("📋 购买总次数",  f"{total_batches} 次",              "#4A4A2A"),
            ("🎉 合计中奖",    f"{total_wins} 注",                 "#6A3A2A"),
            ("📈 平均每%s花费" % period_name,
             f"¥ {total_cost/period_count:.1f}" if period_count else "¥ 0",
             "#2A4A4A"),
        ]
        for title, val, bg in card_info:
            c = tk.Frame(summary_frame, bg=bg, padx=12, pady=8,
                         highlightbackground="#2A4A2A", highlightthickness=1)
            c.pack(side="left", padx=5, pady=2)
            tk.Label(c, text=title, bg=bg, fg="#AADDAA",
                     font=("Microsoft YaHei", 8)).pack()
            tk.Label(c, text=val, bg=bg, fg="#FFFFFF",
                     font=("Microsoft YaHei", 12, "bold")).pack()

        # ── 图形化花费趋势（文字柱状图）──────────────────────
        bar_frame = tk.LabelFrame(parent, text=f"  花费趋势（最近 {min(len(data), 12)} 个{period_name}）  ",
                                  bg=BG, fg="#88FF88",
                                  font=("Microsoft YaHei", 9, "bold"),
                                  padx=12, pady=8)
        bar_frame.pack(fill="x", padx=12, pady=(4, 4))

        recent = data[:12][::-1]   # 最近 12 个，正序排列
        if recent:
            max_cost = max(r["cost"] for r in recent) or 1
            bar_width = 30   # 每根柱子最大字符数

            for row in recent:
                # 标签
                label_key = list(keys)[0]
                label_val = row[label_key]
                pct = row["cost"] / max_cost
                bar_len = max(1, int(pct * bar_width))
                bar_str = "█" * bar_len

                # 颜色：花费越高越红
                if pct >= 0.8:
                    bar_color = "#FF4444"
                elif pct >= 0.5:
                    bar_color = "#FFAA00"
                else:
                    bar_color = "#44CC44"

                row_f = tk.Frame(bar_frame, bg=BG)
                row_f.pack(fill="x", pady=1)

                tk.Label(row_f, text=f"{label_val:>10}",
                         bg=BG, fg="#AADDAA",
                         font=("Consolas", 9)).pack(side="left")
                tk.Label(row_f, text=bar_str,
                         bg=BG, fg=bar_color,
                         font=("Consolas", 9)).pack(side="left", padx=4)
                tk.Label(row_f, text=f"¥{row['cost']:.0f}  {row['batches']}次"
                                     + (f"  中{row['win_tickets']}注" if row["win_tickets"] > 0 else ""),
                         bg=BG, fg="#CCFFCC",
                         font=("Microsoft YaHei", 9)).pack(side="left")

        # ── 明细表格 ───────────────────────────────────────────
        table_frame = tk.Frame(parent, bg=BG)
        table_frame.pack(fill="both", expand=True, padx=12, pady=(4, 8))

        tree_style = f"LedgerPeriod{period_name}.Treeview"
        s = ttk.Style()
        s.configure(tree_style,
                    background="#1A2A1A", fieldbackground="#1A2A1A",
                    rowheight=26, font=("Microsoft YaHei", 9),
                    foreground="#CCFFCC")
        s.configure(f"{tree_style}.Heading",
                    background="#2A5A2A", foreground="white",
                    font=("Microsoft YaHei", 9, "bold"), relief="flat")
        s.map(tree_style,
              background=[("selected", "#2A6A2A")],
              foreground=[("selected", "#FFFFFF")])

        tree = ttk.Treeview(table_frame, columns=list(keys),
                            show="headings", selectmode="browse",
                            style=tree_style)
        for key, header, width in zip(keys, headers, widths):
            tree.heading(key, text=header, anchor="center")
            tree.column(key, width=width, anchor="center", minwidth=60)

        tree.tag_configure("high",   background="#2A1A1A", foreground="#FFAAAA")
        tree.tag_configure("medium", background="#2A2A1A", foreground="#FFFFAA")
        tree.tag_configure("low",    background="#1A2A1A", foreground="#AAFFAA")
        tree.tag_configure("win",    background="#1A3A1A", foreground="#00FF88")

        for row in data:
            vals = tuple(
                f"¥{row[k]:.0f}" if k in ("cost",) else row[k]
                for k in keys
            )
            cost_v = row["cost"]
            if cost_v == 0:
                tag = "low"
            elif row["win_tickets"] > 0:
                tag = "win"
            elif cost_v >= 50:
                tag = "high"
            elif cost_v >= 20:
                tag = "medium"
            else:
                tag = "low"
            tree.insert("", "end", values=vals, tags=(tag,))

        ysb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=ysb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)




# ─── 程序入口 ────────────────────────────────────────────────
if __name__ == "__main__":
    # 防止 worker 线程重复启动 GUI（APP_LAUNCHED 标志由主进程在启动 GUI 前设置）
    if os.environ.get("APP_LAUNCHED") == "1":
        os.chdir(BASE_DIR)
        init_db()
    else:
        os.environ["APP_LAUNCHED"] = "1"
        os.chdir(BASE_DIR)
        init_db()  # 初始化数据库表
        app = App()
        app.mainloop()
